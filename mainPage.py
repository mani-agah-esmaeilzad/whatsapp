import os
import time
import threading
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import customtkinter as ctk
import openpyxl
import requests
import socket

# Selenium Imports
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ============================
# زبان‌ها و ترجمه‌ها
# ============================
languages = {
    "English": {
        "title": "WhatsApp Marketing",
        "send_button": "Send",
        "connection_status_default": "Connection Status: Unknown",
        "error_connection": "Connection error. Please check your internet or WhatsApp login.",
        "error_empty": "No phone, message, or poll to send.",
        "success_sent": "Message sent successfully!",
        "error_sending": "Error sending message:",
        "exit_button": "Exit",
        "upload_excel": "Upload Excel",
        "download_sample_excel": "Download Sample Excel",
        "target_label": "Targets",
        "number_column": "Number",
        "name_column": "Name",
        "message_section": "Message Section",
        "message_tab": "Message",
        "polls": "Polls",
        "add_poll": "Add Poll",
        "buttons": "Buttons",
        "add_button": "Add Button",
        "attachments": "Attachments",
        "add_file": "Add File",
        "delay_settings": "Delay Settings",
        "wait_label": "Wait",
        "seconds_after_every": "seconds after every",
        "messages_label": "messages",
        "send_mode": "Send Mode",
        "manual": "Manual",
        "auto": "Auto",
        "group_send": "Group Send",
        "single_send": "Single Send",
        "phone_label": "Phone:",
        "single_message_label": "Single Message:",
        "language_option": "English",
        "attachment_type": "Attachment Type",
        "image": "Image",
        "sticker": "Sticker",
        "mp3": "MP3",
        "video": "Video",
        "run_page_title": "Single Message Run",
        "initiate_qr": "Initiate WhatsApp & Scan QR Code from your mobile",
        "click_to_initiate": "CLICK TO INITIATE",
        "start_sending": "START SENDING",
        "single_sent_success": "Message sent successfully!",
        "confirm_send": "Proceed to send?"
    },
    "Persian": {
        "title": "واتساپ مارکتینگ",
        "send_button": "ارسال",
        "connection_status_default": "وضعیت اتصال: نامشخص",
        "error_connection": "خطا در اتصال. لطفاً اینترنت یا ورود به واتساپ را بررسی کنید.",
        "error_empty": "شماره، پیام یا نظرسنجی برای ارسال وجود ندارد.",
        "success_sent": "پیام با موفقیت ارسال شد!",
        "error_sending": "خطا در ارسال پیام:",
        "exit_button": "خروج",
        "upload_excel": "بارگذاری اکسل",
        "download_sample_excel": "دانلود نمونه اکسل",
        "target_label": "مخاطبان",
        "number_column": "شماره",
        "name_column": "نام",
        "message_section": "بخش پیام",
        "message_tab": "پیام",
        "polls": "نظرسنجی‌ها",
        "add_poll": "افزودن نظرسنجی",
        "buttons": "دکمه‌ها",
        "add_button": "افزودن دکمه",
        "attachments": "پیوست‌ها",
        "add_file": "افزودن فایل",
        "delay_settings": "تنظیمات تأخیر",
        "wait_label": "انتظار",
        "seconds_after_every": "ثانیه بعد از هر",
        "messages_label": "پیام",
        "send_mode": "وضعیت ارسال",
        "manual": "دستی",
        "auto": "خودکار",
        "group_send": "ارسال گروهی",
        "single_send": "ارسال تکی",
        "phone_label": "شماره:",
        "single_message_label": "پیام تکی:",
        "language_option": "فارسی",
        "attachment_type": "نوع پیوست",
        "image": "عکس",
        "sticker": "استیکر",
        "mp3": "MP3",
        "video": "ویدیو",
        "run_page_title": "ارسال پیام تکی",
        "initiate_qr": "واتساپ را راه‌اندازی کرده و کد QR را از گوشی اسکن کنید",
        "click_to_initiate": "برای شروع کلیک کنید",
        "start_sending": "شروع ارسال",
        "single_sent_success": "پیام با موفقیت ارسال شد!",
        "confirm_send": "آیا از ارسال پیام مطمئن هستید؟"
    },
    "Arabic": {
        "title": "تسويق واتساب",
        "send_button": "إرسال",
        "connection_status_default": "حالة الاتصال: غير محددة",
        "error_connection": "خطأ في الاتصال. يرجى التحقق من الإنترنت أو تسجيل الدخول إلى واتساپ.",
        "error_empty": "لا يوجد رقم أو رسالة أو استطلاع للإرسال.",
        "success_sent": "تم إرسال الرسالة بنجاح!",
        "error_sending": "خطأ في إرسال الرسالة:",
        "exit_button": "خروج",
        "upload_excel": "تحميل إكسل",
        "download_sample_excel": "تحميل عينة إكسل",
        "target_label": "المستهدفون",
        "number_column": "رقم",
        "name_column": "اسم",
        "message_section": "قسم الرسالة",
        "message_tab": "رسالة",
        "polls": "الاستطلاعات",
        "add_poll": "إضافة استطلاع",
        "buttons": "الأزرار",
        "add_button": "إضافة زر",
        "attachments": "المرفقات",
        "add_file": "إضافة ملف",
        "delay_settings": "إعدادات التأخير",
        "wait_label": "انتظار",
        "seconds_after_every": "ثانية بعد كل",
        "messages_label": "رسالة",
        "send_mode": "وضع الإرسال",
        "manual": "يدوي",
        "auto": "تلقائي",
        "group_send": "إرسال جماعي",
        "single_send": "إرسال فردي",
        "phone_label": "رقم الهاتف:",
        "single_message_label": "رسالة فردية:",
        "language_option": "العربية",
        "attachment_type": "نوع المرفق",
        "image": "صورة",
        "sticker": "ملصق",
        "mp3": "MP3",
        "video": "فيديو",
        "run_page_title": "تشغيل رسالة فردية",
        "initiate_qr": "ابدأ واتساب وامسح رمز الاستجابة السريعة من هاتفك",
        "click_to_initiate": "انقر للبدء",
        "start_sending": "ابدأ الإرسال",
        "single_sent_success": "تم إرسال الرسالة بنجاح!",
        "confirm_send": "هل أنت متأكد من الإرسال؟"
    }
}

current_lang = "English"

def get_qr_confirmation_message():
    if current_lang == "Persian":
        return "آیا بارکد را اسکن کردید؟"
    elif current_lang == "Arabic":
        return "هل قمت بمسح رمز الاستجابة السريعة؟"
    else:
        return "Did you scan the QR code?"

def get_font(size, weight="normal"):
    if current_lang == "Persian":
        return ctk.CTkFont(family="vazir", size=size, weight=weight)
    elif current_lang == "Arabic":
        return ctk.CTkFont(family="Arial", size=size, weight=weight)
    else:
        return ctk.CTkFont(family="Segoe UI", size=size, weight=weight)

def get_ttk_font():
    if current_lang == "Persian":
        return ("vazir", 12)
    elif current_lang == "Arabic":
        return ("Arial", 12)
    else:
        return ("Segoe UI", 12)

def get_ttk_heading_font():
    if current_lang == "Persian":
        return ("vazir", 12, "bold")
    elif current_lang == "Arabic":
        return ("Arial", 12, "bold")
    else:
        return ("Segoe UI", 12, "bold")

# ============================
# دریافت آدرس آیپی سیستم
# ============================
def get_local_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
    except Exception:
        ip = "127.0.0.1"
    finally:
        s.close()
    return ip

# ============================
# ارسال لاگ به API دات نت
# ============================
def send_log_to_api(phone, status, message_text=""):
    if status == "Sending...":
        return
    ip = get_local_ip()
    url = "https://maniagah.ir/api/logs"  # آدرس API دات نت (تنظیم به نیاز شما)
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    data = {
        "Phone": phone,
        "Status": status,
        "Platform": "application",
        "Text": message_text,
        "SystemIp": ip,
        "subscriptionType": "Free"
    }
    try:
        response = requests.post(url, json=data, headers=headers, verify=False)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"Error sending log to API: {e}")

# ============================
# ویجت جایگزین Spinbox
# ============================
class SpinboxAlternative(ctk.CTkFrame):
    def __init__(self, parent, from_=1, to=9999, width=60, textvariable=None, **kwargs):
        # Remove unsupported arguments from kwargs
        button_color = kwargs.pop('button_color', '#e94560')
        button_hover_color = kwargs.pop('button_hover_color', '#b81d3d')
        
        super().__init__(parent, **kwargs)
        self.from_ = from_
        self.to = to
        if textvariable is None:
            self.var = tk.IntVar(value=self.from_)
        else:
            self.var = textvariable
            
        # Configure buttons with proper styling
        self.btn_minus = ctk.CTkButton(
            self, 
            text="-", 
            width=30, 
            command=self.decrement, 
            corner_radius=5,
            fg_color=button_color,
            hover_color=button_hover_color
        )
        self.btn_minus.grid(row=0, column=0, padx=2)
        
        self.entry = ctk.CTkEntry(self, width=width, textvariable=self.var)
        self.entry.grid(row=0, column=1, padx=2)
        
        self.btn_plus = ctk.CTkButton(
            self, 
            text="+", 
            width=30, 
            command=self.increment, 
            corner_radius=5,
            fg_color=button_color,
            hover_color=button_hover_color
        )
        self.btn_plus.grid(row=0, column=2, padx=2)
    
    def increment(self):
        try:
            value = int(self.var.get())
        except ValueError:
            value = self.from_
        if value < self.to:
            value += 1
        self.var.set(value)
    
    def decrement(self):
        try:
            value = int(self.var.get())
        except ValueError:
            value = self.from_
        if value > self.from_:
            value -= 1
        self.var.set(value)

# ============================
# پنجره افزودن نظرسنجی
# ============================
class PollWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        if current_lang == "Persian":
            self.title("افزودن نظرسنجی")
        elif current_lang == "Arabic":
            self.title("إضافة استطلاع")
        else:
            self.title("Add Poll")
        self.geometry("400x400")
        question_label = ctk.CTkLabel(self, text=("سوال نظرسنجی:" if current_lang=="Persian" else ("سؤال الاستطلاع:" if current_lang=="Arabic" else "Poll Question:")), font=get_font(14, "bold"))
        question_label.pack(padx=10, pady=10)
        self.question_entry = ctk.CTkEntry(self, font=get_font(14))
        self.question_entry.pack(padx=10, pady=10, fill="x")
        options_label = ctk.CTkLabel(self, text=("گزینه‌های نظرسنجی (هر گزینه در یک خط):" if current_lang=="Persian" else ("خيارات الاستطلاع (سطر لكل خيار):" if current_lang=="Arabic" else "Poll Options (one per line):")), font=get_font(14, "bold"))
        options_label.pack(padx=10, pady=10)
        self.options_textbox = ctk.CTkTextbox(self, font=get_font(14), height=150, wrap="word")
        self.options_textbox.pack(padx=10, pady=10, fill="both", expand=True)
        add_button = ctk.CTkButton(self, text=("افزودن نظرسنجی" if current_lang=="Persian" else ("إضافة استطلاع" if current_lang=="Arabic" else "Add Poll")), command=self.add_poll, fg_color="#0078AA", corner_radius=5, font=get_font(14, "bold"))
        add_button.pack(padx=10, pady=10)
    
    def add_poll(self):
        question = self.question_entry.get().strip()
        options_text = self.options_textbox.get("1.0", "end").strip()
        if not question or not options_text:
            messagebox.showerror("Error" if current_lang=="English" else ("خطا" if current_lang=="Persian" else "خطأ"),
                                 "Please enter both question and options." if current_lang=="English" else ("لطفاً سوال و گزینه‌ها را وارد کنید." if current_lang=="Persian" else "يرجى إدخال السؤال والخيارات."))
            return
        options = [opt.strip() for opt in options_text.splitlines() if opt.strip()]
        if len(options) < 2:
            messagebox.showerror("Error" if current_lang=="English" else ("خطا" if current_lang=="Persian" else "خطأ"),
                                 "Please enter at least two options." if current_lang=="English" else ("لطفاً حداقل دو گزینه وارد کنید." if current_lang=="Persian" else "يرجى إدخال خيارين على الأقل."))
            return
        poll_data = {"question": question, "options": options}
        self.master.polls_list.append(poll_data)
        messagebox.showinfo("Info" if current_lang=="English" else ("اطلاع" if current_lang=="Persian" else "معلومات"),
                            "Poll added successfully." if current_lang=="English" else ("نظرسنجی با موفقیت افزوده شد." if current_lang=="Persian" else "تم إضافة الاستطلاع بنجاح."))
        self.destroy()

# ============================
# پنجره فرآیند ارسال پیام گروهی
# ============================
class SendProcessWindow(ctk.CTkToplevel):
    def __init__(self, parent, numbers_list, messages, attachments, polls, delay_seconds, delay_every_msgs, mode):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        self.title("Run")
        self.geometry("900x550")
        self.numbers_list = numbers_list
        self.messages = messages
        self.attachments = attachments
        self.polls = polls
        self.delay_seconds = delay_seconds
        self.delay_every_msgs = delay_every_msgs
        self.mode = mode
        self.paused = False
        self.stopped = False
        self.driver = None
        self.scan_confirmed = None
        
        # --- Top Frame: Initiate Connection ---
        top_frame = ctk.CTkFrame(self, fg_color="#1f2833", corner_radius=10)
        top_frame.pack(fill="x", pady=10, padx=10)
        initiate_label = ctk.CTkLabel(top_frame, text="Initiate WhatsApp & Scan QR Code from your mobile", font=get_font(14, "bold"), text_color="white")
        initiate_label.pack(side="left", padx=10, pady=10)
        self.status_label = ctk.CTkLabel(top_frame, text=languages[current_lang]["connection_status_default"], font=get_font(14), text_color="white")
        self.status_label.pack(side="right", padx=10)
        self.initiate_button = ctk.CTkButton(top_frame, text=languages[current_lang]["click_to_initiate"], fg_color="#0078AA", command=self.initiate_connection, corner_radius=5)
        self.initiate_button.pack(side="left", padx=10)
        
        # --- Middle Frame: Control Buttons & Progress ---
        middle_frame = ctk.CTkFrame(self, fg_color="#20232a", corner_radius=10)
        middle_frame.pack(fill="x", pady=10, padx=10)
        self.start_button = ctk.CTkButton(middle_frame, text="START", fg_color="#4CAF50", command=self.start_sending, corner_radius=5, font=get_font(14, "bold"))
        self.start_button.pack(side="left", padx=5)
        self.pause_button = ctk.CTkButton(middle_frame, text="PAUSE", fg_color="#FF9800", command=self.pause_sending, corner_radius=5, font=get_font(14, "bold"))
        self.pause_button.pack(side="left", padx=5)
        self.stop_button = ctk.CTkButton(middle_frame, text="STOP", fg_color="#F44336", command=self.stop_sending, corner_radius=5, font=get_font(14, "bold"))
        self.stop_button.pack(side="left", padx=5)
        self.progress_label = ctk.CTkLabel(middle_frame, text="0% Completed [0/{}]".format(len(self.numbers_list)), font=get_font(14), text_color="white")
        self.progress_label.pack(side="right", padx=10)
        
        # --- Log Frame ---
        log_frame = ctk.CTkFrame(self, fg_color="#20232a", corner_radius=10)
        log_frame.pack(fill="both", expand=True, padx=10, pady=10)
        self.log_tree = ttk.Treeview(log_frame, columns=("chat_name", "status"), show="headings", style="Custom.Treeview")
        self.log_tree.heading("chat_name", text="Chat Name")
        self.log_tree.heading("status", text="Status")
        self.log_tree.column("chat_name", width=300)
        self.log_tree.column("status", width=200)
        self.log_tree.pack(fill="both", expand=True, padx=10, pady=10)
        
        # --- Notes Frame ---
        notes_frame = ctk.CTkFrame(self, fg_color="#1f2833", corner_radius=10)
        notes_frame.pack(fill="x", padx=10, pady=10)
        notes_label = ctk.CTkLabel(notes_frame, text=("Important Notes:\n1) Make sure your WhatsApp Web is logged in.\n2) Keep your phone connected to the internet.\n3) Do not close the browser tab during sending.\n"), font=get_font(12), justify="left", text_color="white")
        notes_label.pack(side="left", padx=10, pady=10)
    
    def initiate_connection(self):
        threading.Thread(target=self._initiate_connection_worker, daemon=True).start()
    
    def _initiate_connection_worker(self):
        try:
            self.driver = webdriver.Chrome()
            self.driver.get("https://web.whatsapp.com")
            self.after(0, lambda: self.status_label.configure(text="Status: Waiting for login"))
            WebDriverWait(self.driver, 60).until(
                EC.presence_of_element_located((By.XPATH, "//div[@id='pane-side']"))
            )
            self.after(0, lambda: self.status_label.configure(text="Status: Connected"))
            self.scan_confirmed = None
            self.after(0, self.prompt_scan_confirmation)
            while self.scan_confirmed is None:
                time.sleep(0.1)
            if self.scan_confirmed:
                self.driver.set_window_position(-2000, 0)
            else:
                self.after(0, lambda: messagebox.showwarning("Warning", "Please scan the QR code before proceeding."))
        except Exception as e:
            self.after(0, lambda: self.status_label.configure(text="Status: Disconnected"))
            self.after(0, lambda: messagebox.showerror("Error", f"{languages[current_lang]['error_connection']} {e}"))
    
    def prompt_scan_confirmation(self):
        result = messagebox.askyesno("QR Code Scanned", get_qr_confirmation_message())
        self.scan_confirmed = result
    
    def stop_sending(self):
        self.stopped = True
    
    def pause_sending(self):
        self.paused = True
    
    def start_sending(self):
        if not self.driver:
            messagebox.showerror("Error", "Please initiate connection first.")
            return
        self.paused = False
        self.stopped = False
        self.start_button.configure(state="disabled")
        threading.Thread(target=self.send_messages, daemon=True).start()
    
    def send_messages(self):
        total = len(self.numbers_list)
        for idx, contact in enumerate(self.numbers_list, start=1):
            if self.stopped:
                break
            while self.paused and not self.stopped:
                time.sleep(0.5)
            phone = contact["number"]
            self.update_log(phone, "Sending...")
            try:
                self.driver.get(f"https://web.whatsapp.com/send?phone={phone}")
                time.sleep(3)
                if self.attachments:
                    att_type = self.master.attachment_type.get() if hasattr(self.master, "attachment_type") else languages[current_lang]["image"]
                    if att_type == languages[current_lang]["image"]:
                        try:
                            attach_button = WebDriverWait(self.driver, 30).until(
                                EC.element_to_be_clickable((By.XPATH, "//button[@title='Attach']"))
                            )
                            attach_button.click()
                            file_input = WebDriverWait(self.driver, 30).until(
                                EC.presence_of_element_located((By.XPATH, "//input[@accept='image/*']"))
                            )
                            file_input.send_keys(self.attachments[0])
                            time.sleep(5)
                            caption = " ".join(self.messages) if self.messages else " "
                            caption_box = WebDriverWait(self.driver, 10).until(
                                EC.visibility_of_element_located((By.XPATH, "//div[@contenteditable='true'][@data-tab='10']"))
                            )
                            caption_box.clear()
                            caption_box.send_keys(caption)
                            send_button = WebDriverWait(self.driver, 30).until(
                                EC.element_to_be_clickable((By.XPATH, "//button[contains(@aria-label, 'Send photo')]"))
                            )
                            self.driver.execute_script("arguments[0].click();", send_button)
                            ActionChains(self.driver).send_keys(Keys.ENTER).perform()
                            time.sleep(2)
                        except Exception as e:
                            self.update_log(phone, f"Image send failed: {e}")
                    elif att_type == languages[current_lang]["sticker"]:
                        try:
                            attach_button = WebDriverWait(self.driver, 30).until(
                                EC.element_to_be_clickable((By.XPATH, "//button[@title='Attach']"))
                            )
                            attach_button.click()
                            file_input = WebDriverWait(self.driver, 30).until(
                                EC.presence_of_element_located((By.XPATH, "//input[@accept='image/*']"))
                            )
                            file_input.send_keys(self.attachments[0])
                            time.sleep(5)
                            send_button = WebDriverWait(self.driver, 30).until(
                                EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Send'] | //span[@data-icon='send']"))
                            )
                            send_button.click()
                            time.sleep(2)
                        except Exception as e:
                            self.update_log(phone, f"Sticker send failed: {e}")
                    elif att_type == languages[current_lang]["mp3"]:
                        try:
                            attach_button = WebDriverWait(self.driver, 30).until(
                                EC.element_to_be_clickable((By.XPATH, "//button[@title='Attach']"))
                            )
                            attach_button.click()
                            file_input = WebDriverWait(self.driver, 30).until(
                                EC.presence_of_element_located((By.XPATH, "//input[@accept='audio/*']"))
                            )
                            file_input.send_keys(self.attachments[0])
                            time.sleep(5)
                            send_button = WebDriverWait(self.driver, 30).until(
                                EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Send'] | //span[@data-icon='send']"))
                            )
                            send_button.click()
                            time.sleep(2)
                        except Exception as e:
                            self.update_log(phone, f"MP3 send failed: {e}")
                    elif att_type == languages[current_lang]["video"]:
                        try:
                            attach_button = WebDriverWait(self.driver, 30).until(
                                EC.element_to_be_clickable((By.XPATH, "//button[@title='Attach']"))
                            )
                            attach_button.click()
                            file_input = WebDriverWait(self.driver, 30).until(
                                EC.presence_of_element_located((By.XPATH, "//input[@accept='image/*,video/mp4,video/3gpp,video/quicktime']"))
                            )
                            file_input.send_keys(self.attachments[0])
                            time.sleep(5)
                            caption = " ".join(self.messages) if self.messages else " "
                            caption_box = WebDriverWait(self.driver, 10).until(
                                EC.visibility_of_element_located((By.XPATH, "//div[@contenteditable='true'][@data-tab='10']"))
                            )
                            caption_box.clear()
                            caption_box.send_keys(caption)
                            send_button = WebDriverWait(self.driver, 30).until(
                                EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Send'] | //span[@data-icon='send']"))
                            )
                            send_button.click()
                            time.sleep(2)
                        except Exception as e:
                            self.update_log(phone, f"Video send failed: {e}")
                else:
                    message_box = WebDriverWait(self.driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, "//footer//div[@contenteditable='true']"))
                    )
                    message_box.click()
                    time.sleep(0.5)
                    # Combine all messages with line breaks
                    combined_message = "\n".join(self.messages)
                    # Replace placeholders in the combined message
                    personalized_msg = combined_message.replace("{NAME}", contact.get("name", "")).replace("{NUMBER}", contact.get("number", ""))
                    # Split into lines and send with proper line breaks
                    lines = personalized_msg.split('\n')
                    for i, line in enumerate(lines):
                        message_box.send_keys(line)
                        if i < len(lines) - 1:  # If not the last line
                            ActionChains(self.driver).key_down(Keys.SHIFT).send_keys(Keys.ENTER).key_up(Keys.SHIFT).perform()
                            time.sleep(0.1)
                    # Send the final message
                    message_box.send_keys(Keys.ENTER)
                    time.sleep(1)
                self.update_log(phone, "Success")
            except Exception as e:
                self.update_log(phone, f"Failed: {e}")
            percent = int((idx / total) * 100)
            self.progress_label.configure(text=f"{percent}% Completed [{idx}/{total}]")
            if (idx % self.delay_every_msgs == 0) and (idx < total):
                time.sleep(self.delay_seconds)
    
    def update_log(self, chat_name, status):
        self.log_tree.insert("", tk.END, values=(chat_name, status))
        self.log_tree.update_idletasks()
        if status != "Sending...":
            message_text = " ".join(self.messages) if self.messages else ""
            send_log_to_api(chat_name, status, message_text)

# ============================
# تابع ترجمه متن (با استفاده از API گوگل)
# ============================
def translate_text(text, target_lang):
    try:
        url = "https://translate.googleapis.com/translate_a/single"
        params = {
            "client": "gtx",
            "sl": "auto",
            "tl": target_lang,
            "dt": "t",
            "q": text
        }
        response = requests.get(url, params=params)
        if response.status_code == 200:
            result = response.json()
            translated_text = ''.join([segment[0] for segment in result[0] if segment[0]])
            return translated_text
        else:
            return None
    except Exception as e:
        print("Translation error:", e)
        return None

# ============================
# پنجره ارسال پیام تکی (ورودی)
# ============================
class SingleMessageInputWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        if current_lang == "Persian":
            self.title("ارسال پیام تکی")
        elif current_lang == "Arabic":
            self.title("إرسال رسالة فردية")
        else:
            self.title("Send Single Message")
        self.geometry("500x500")
        self.attachment_path = None
        
        # ورودی شماره تلفن
        phone_frame = ctk.CTkFrame(self, fg_color="#1f2833", corner_radius=10)
        phone_frame.pack(fill="x", padx=10, pady=10)
        phone_label = ctk.CTkLabel(phone_frame, text=languages[current_lang]["phone_label"], font=get_font(14, "bold"), text_color="white")
        phone_label.pack(side="left", padx=10, pady=10)
        self.phone_entry = ctk.CTkEntry(phone_frame, font=get_font(14))
        self.phone_entry.pack(side="left", fill="x", expand=True, padx=10, pady=10)
        
        # ورودی پیام
        message_frame = ctk.CTkFrame(self, fg_color="#1f2833", corner_radius=10)
        message_frame.pack(fill="both", expand=True, padx=10, pady=10)
        message_label = ctk.CTkLabel(message_frame, text=languages[current_lang]["single_message_label"], font=get_font(14, "bold"), text_color="white")
        message_label.pack(anchor="nw", padx=10, pady=10)
        self.message_textbox = ctk.CTkTextbox(message_frame, font=get_font(14), wrap="word")
        self.message_textbox.pack(fill="both", expand=True, padx=10, pady=10)
        
        # دکمه افزودن پیوست
        attach_button = ctk.CTkButton(self, text=languages[current_lang]["add_file"], command=self.attach_file, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        attach_button.pack(padx=10, pady=5)
        
        # دکمه‌های ترجمه برای پیام تکی
        if current_lang == "Persian":
            translate_arabic_text = "ترجمه به عربی"
            translate_english_text = "ترجمه به انگلیسی"
        elif current_lang == "Arabic":
            translate_arabic_text = "ترجمة إلى العربية"
            translate_english_text = "ترجمة إلى الإنجليزية"
        else:
            translate_arabic_text = "Translate to Arabic"
            translate_english_text = "Translate to English"
            
        translate_arabic_button = ctk.CTkButton(self, text=translate_arabic_text, command=self.translate_to_arabic, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        translate_arabic_button.pack(padx=10, pady=5)
        
        translate_english_button = ctk.CTkButton(self, text=translate_english_text, command=self.translate_to_english, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        translate_english_button.pack(padx=10, pady=5)
        
        # دکمه برای رفتن به صفحه ران ارسال پیام تکی
        proceed_button = ctk.CTkButton(self, text=languages[current_lang]["send_button"], command=self.proceed_to_run, fg_color="#4CAF50", corner_radius=5, font=get_font(14, "bold"))
        proceed_button.pack(padx=10, pady=10)
    
    def attach_file(self):
        file_path = filedialog.askopenfilename(title=languages[current_lang]["add_file"], filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.gif"), ("Audio Files", "*.mp3"), ("Video Files", "*.mp4;*.avi;*.mov")])
        if file_path:
            self.attachment_path = file_path
            messagebox.showinfo("Info", f"Attached: {os.path.basename(file_path)}")
    
    def proceed_to_run(self):
        phone = self.phone_entry.get().strip()
        message = self.message_textbox.get("1.0", "end").strip()
        if not phone or not message:
            messagebox.showerror("Error", languages[current_lang]["error_empty"])
            return
        corrected_phone = self.master.correct_phone_number(phone)
        if messagebox.askyesno("Confirm", languages[current_lang]["confirm_send"]):
            run_window = SingleMessageRunWindow(self.master, corrected_phone, message, self.attachment_path)
            self.destroy()
    
    def translate_to_arabic(self):
        original_text = self.message_textbox.get("1.0", "end").strip()
        if not original_text:
            messagebox.showwarning("Warning", "No text to translate.")
            return
        translated = translate_text(original_text, "ar")
        if translated:
            self.message_textbox.delete("1.0", "end")
            self.message_textbox.insert("end", translated)
        else:
            messagebox.showerror("Error", "Translation to Arabic failed.")
    
    def translate_to_english(self):
        original_text = self.message_textbox.get("1.0", "end").strip()
        if not original_text:
            messagebox.showwarning("Warning", "No text to translate.")
            return
        translated = translate_text(original_text, "en")
        if translated:
            self.message_textbox.delete("1.0", "end")
            self.message_textbox.insert("end", translated)
        else:
            messagebox.showerror("Error", "Translation to English failed.")

# ============================
# پنجره ارسال پیام تکی (صفحه اجرا)
# ============================
class SingleMessageRunWindow(ctk.CTkToplevel):
    def __init__(self, parent, phone, message, attachment):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        self.title(languages[current_lang]["run_page_title"])
        self.geometry("700x500")
        self.phone = phone
        self.message = message
        self.attachment = attachment 
        self.driver = None
        self.scan_confirmed = None
        
        top_frame = ctk.CTkFrame(self, fg_color="#1f2833", corner_radius=10)
        top_frame.pack(fill="x", pady=10, padx=10)
        initiate_label = ctk.CTkLabel(top_frame, text=languages[current_lang]["initiate_qr"], font=get_font(14, "bold"), text_color="white")
        initiate_label.pack(side="left", padx=10, pady=10)
        self.status_label = ctk.CTkLabel(top_frame, text=languages[current_lang]["connection_status_default"], font=get_font(14), text_color="white")
        self.status_label.pack(side="right", padx=10)
        self.initiate_button = ctk.CTkButton(top_frame, text=languages[current_lang]["click_to_initiate"], fg_color="#0078AA", command=self.initiate_connection, corner_radius=5)
        self.initiate_button.pack(side="left", padx=10)
        
        # --- Middle Frame: نمایش اطلاعات پیام ---
        middle_frame = ctk.CTkFrame(self, fg_color="#20232a", corner_radius=10)
        middle_frame.pack(fill="both", expand=True, pady=10, padx=10)
        info_label = ctk.CTkLabel(middle_frame, text=f"Phone: {self.phone}\nMessage:\n{self.message}", font=get_font(14), text_color="white", justify="left")
        info_label.pack(padx=10, pady=10, anchor="w")
        
        # --- Control Frame ---
        control_frame = ctk.CTkFrame(self, fg_color="#20232a", corner_radius=10)
        control_frame.pack(fill="x", pady=10, padx=10)
        self.start_button = ctk.CTkButton(control_frame, text=languages[current_lang]["start_sending"], fg_color="#4CAF50", command=self.start_sending, corner_radius=5, font=get_font(14, "bold"))
        self.start_button.pack(side="left", padx=10, pady=5)
        exit_button = ctk.CTkButton(control_frame, text=languages[current_lang]["exit_button"], command=self.on_close, fg_color="#F44336", corner_radius=5, font=get_font(14))
        exit_button.pack(side="right", padx=10, pady=5)
        
        # --- Log Frame ---
        log_frame = ctk.CTkFrame(self, fg_color="#20232a", corner_radius=10)
        log_frame.pack(fill="both", expand=True, padx=10, pady=10)
        self.log_tree = ttk.Treeview(log_frame, columns=("status",), show="headings", style="Custom.Treeview")
        self.log_tree.heading("status", text="Status")
        self.log_tree.column("status", width=600)
        self.log_tree.pack(fill="both", expand=True, padx=10, pady=10)
    
    def initiate_connection(self):
        threading.Thread(target=self._initiate_connection_worker, daemon=True).start()
    
    def _initiate_connection_worker(self):
        try:
            self.driver = webdriver.Chrome()
            self.driver.get("https://web.whatsapp.com")
            self.after(0, lambda: self.status_label.configure(text="Status: Waiting for login"))
            WebDriverWait(self.driver, 60).until(
                EC.presence_of_element_located((By.XPATH, "//div[@id='pane-side']"))
            )
            self.after(0, lambda: self.status_label.configure(text="Status: Connected"))
            self.scan_confirmed = None
            self.after(0, self.prompt_scan_confirmation)
            while self.scan_confirmed is None:
                time.sleep(0.1)
            if self.scan_confirmed:
                self.driver.set_window_position(-2000, 0)
            else:
                self.after(0, lambda: messagebox.showwarning("Warning", "Please scan the QR code before proceeding."))
        except Exception as e:
            self.after(0, lambda: self.status_label.configure(text="Status: Disconnected"))
            self.after(0, lambda: messagebox.showerror("Error", f"{languages[current_lang]['error_connection']} {e}"))
    
    def prompt_scan_confirmation(self):
        result = messagebox.askyesno("QR Code Scanned", get_qr_confirmation_message())
        self.scan_confirmed = result
    
    def start_sending(self):
        if not self.driver:
            messagebox.showerror("Error", "Please initiate connection first.")
            return
        threading.Thread(target=self.send_message, daemon=True).start()
    
    def send_message(self):
        try:
            self.driver.get(f"https://web.whatsapp.com/send?phone={self.phone}")
            time.sleep(3)
            if self.attachment:
                att_type = self.master.attachment_type.get() if hasattr(self.master, "attachment_type") else languages[current_lang]["image"]
                if att_type == languages[current_lang]["image"]:
                    try:
                        attach_button = WebDriverWait(self.driver, 30).until(
                            EC.element_to_be_clickable((By.XPATH, "//button[@title='Attach']"))
                        )
                        attach_button.click()
                        file_input = WebDriverWait(self.driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, "//input[@accept='image/*']"))
                        )
                        file_input.send_keys(self.attachment)
                        time.sleep(5)
                        caption_box = WebDriverWait(self.driver, 10).until(
                            EC.visibility_of_element_located((By.XPATH, "//div[@contenteditable='true'][@data-tab='10']"))
                        )
                        caption_box.clear()
                        caption_box.send_keys(self.message)
                        send_button = WebDriverWait(self.driver, 30).until(
                            EC.element_to_be_clickable((By.XPATH, "//button[contains(@aria-label, 'Send photo')]"))
                        )
                        self.driver.execute_script("arguments[0].click();", send_button)
                        ActionChains(self.driver).send_keys(Keys.ENTER).perform()
                        time.sleep(2)
                    except Exception as e:
                        self.update_log(f"Image send failed: {e}")
                        return
                elif att_type == languages[current_lang]["sticker"]:
                    try:
                        attach_button = WebDriverWait(self.driver, 30).until(
                            EC.element_to_be_clickable((By.XPATH, "//button[@title='Attach']"))
                        )
                        attach_button.click()
                        file_input = WebDriverWait(self.driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, "//input[@accept='image/*']"))
                        )
                        file_input.send_keys(self.attachment)
                        time.sleep(5)
                        send_button = WebDriverWait(self.driver, 30).until(
                            EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Send'] | //span[@data-icon='send']"))
                        )
                        send_button.click()
                        time.sleep(2)
                    except Exception as e:
                        self.update_log(f"Sticker send failed: {e}")
                elif att_type == languages[current_lang]["mp3"]:
                    try:
                        attach_button = WebDriverWait(self.driver, 30).until(
                            EC.element_to_be_clickable((By.XPATH, "//button[@title='Attach']"))
                        )
                        attach_button.click()
                        file_input = WebDriverWait(self.driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, "//input[@accept='audio/*']"))
                        )
                        file_input.send_keys(self.attachment)
                        time.sleep(5)
                        send_button = WebDriverWait(self.driver, 30).until(
                            EC.element_to_be_clickable((By.XPATH, "//button[@aria-label='Send'] | //span[@data-icon='send']"))
                        )
                        send_button.click()
                        time.sleep(2)
                    except Exception as e:
                        self.update_log(f"MP3 send failed: {e}")
                elif att_type == languages[current_lang]["video"]:
                    try:
                        attach_button = WebDriverWait(self.driver, 30).until(
                            EC.element_to_be_clickable((By.XPATH, "//button[@title='Attach']"))
                        )
                        attach_button.click()
                        file_input = WebDriverWait(self.driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, "//input[@accept='video/*']"))
                        )
                        file_input.send_keys(self.attachment)
                        time.sleep(5)
                        caption_box = WebDriverWait(self.driver, 10).until(
                            EC.visibility_of_element_located((By.XPATH, "//div[@contenteditable='true'][@data-tab='10']"))
                        )
                        caption_box.clear()
                        caption_box.send_keys(self.message)
                        send_button = WebDriverWait(self.driver, 30).until(
                            EC.element_to_be_clickable((By.XPATH, "//button[contains(@aria-label, 'Send')]"))
                        )
                        send_button.click()
                        time.sleep(2)
                    except Exception as e:
                        self.update_log(f"Video send failed: {e}")
            else:
                message_box = WebDriverWait(self.driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, "//footer//div[@contenteditable='true']"))
                )
                message_box.click()
                time.sleep(0.5)
                # Combine all messages with line breaks
                combined_message = "\n".join(self.messages)
                # Replace placeholders in the combined message
                personalized_msg = combined_message.replace("{NAME}", self.phone).replace("{NUMBER}", self.phone)
                # Split into lines and send with proper line breaks
                lines = personalized_msg.split('\n')
                for i, line in enumerate(lines):
                    message_box.send_keys(line)
                    if i < len(lines) - 1:  # If not the last line
                        ActionChains(self.driver).key_down(Keys.SHIFT).send_keys(Keys.ENTER).key_up(Keys.SHIFT).perform()
                        time.sleep(0.1)
                    # Send the final message
                    message_box.send_keys(Keys.ENTER)
                    time.sleep(1)
                self.update_log(languages[current_lang]["single_sent_success"])
        except Exception as e:
            self.update_log(f"{languages[current_lang]['error_sending']} {e}")
    
    def update_log(self, status):
        self.log_tree.insert("", tk.END, values=(status,))
        self.log_tree.update_idletasks()
        if status != "Sending...":
            send_log_to_api(self.phone, status, self.message)
    
    def on_close(self):
        if self.driver:
            self.driver.quit()
        self.destroy()

# ============================
# کلاس اصلی برنامه
# ============================
class WhatsAppMarketingApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        # Set window to maximized state instead of fullscreen
        self.state("zoomed")
        self.title(languages[current_lang]["title"])
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.numbers_list = []
        self.attachments_list = []
        self.polls_list = []
        self.delay_seconds = tk.IntVar(value=3)
        self.delay_every_msgs = tk.IntVar(value=5)
        self.send_mode_var = tk.StringVar(value="Manual")
        self.attachment_type = tk.StringVar(value=languages[current_lang]["image"])
        self.textboxes = []
        self.driver = None
        
        # Set minimum window size to a more reasonable value
        self.minsize(800, 600)
        
        self.create_ui()
        self.apply_fullscreen_style()
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Custom.Treeview", background="#0b0c10", foreground="white", fieldbackground="#0b0c10", font=get_ttk_font())
        style.configure("Custom.Treeview.Heading", background="#1f2833", foreground="white", font=get_ttk_heading_font())
        style.map("Custom.Treeview", background=[("selected", "#3a3f47")])
    
    def apply_fullscreen_style(self):
        # Get screen dimensions
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        
        # Configure grid weights for better scaling
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        
        # Configure main background color
        self.configure(bg="#1a1a2e")
    
    def create_ui(self):
        # Main background color
        self.configure(bg="#1a1a2e")
        
        # Top Frame with gradient effect and reduced height
        top_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="#16213e", height=60)
        top_frame.pack(side="top", fill="x", padx=10, pady=10)
        
        # Title with reduced font size
        title_label = ctk.CTkLabel(
            top_frame, 
            text=languages[current_lang]["title"], 
            font=get_font(24, "bold"), 
            text_color="#e94560"
        )
        title_label.pack(side="left", padx=20)
        
        # Options frame with modern styling
        options_frame = ctk.CTkFrame(top_frame, corner_radius=0, fg_color="#16213e")
        options_frame.pack(side="right", padx=20)
        
        # Modern styled option menus with reduced size
        appearance_option_menu = ctk.CTkOptionMenu(
            options_frame, 
            values=["Light", "Dark"], 
            command=self.change_appearance_mode, 
            width=120, 
            height=30,
            font=get_font(12),
            fg_color="#1f2937",
            button_color="#e94560",
            button_hover_color="#b81d3d"
        )
        appearance_option_menu.set("Dark")
        appearance_option_menu.pack(side="left", padx=10)
        
        language_option_menu = ctk.CTkOptionMenu(
            options_frame, 
            values=["English", "Persian", "Arabic"], 
            command=self.change_language, 
            width=120, 
            height=30,
            font=get_font(12),
            fg_color="#1f2937",
            button_color="#e94560",
            button_hover_color="#b81d3d"
        )
        language_option_menu.set(languages[current_lang]["language_option"])
        language_option_menu.pack(side="left", padx=10)
        
        # Main content frame with reduced padding
        main_frame = ctk.CTkFrame(self, fg_color="#1f2937", corner_radius=15)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        main_frame.grid_columnconfigure(1, weight=1)
        main_frame.grid_rowconfigure(0, weight=1)
        
        # Left Panel - Targets with reduced width
        left_frame = ctk.CTkFrame(main_frame, corner_radius=15, fg_color="#2d3748", width=250)
        left_frame.grid(row=0, column=0, sticky="nswe", padx=10, pady=10)
        
        # Modern styled target label with reduced font size
        target_label = ctk.CTkLabel(
            left_frame, 
            text=languages[current_lang]["target_label"], 
            font=get_font(20, "bold"), 
            text_color="#e94560"
        )
        target_label.pack(padx=15, pady=15)
        
        # Button frame with reduced padding
        button_frame = ctk.CTkFrame(left_frame, fg_color="#2d3748")
        button_frame.pack(padx=15, pady=10, fill="x")
        
        # Modern styled buttons with reduced size
        upload_button = ctk.CTkButton(
            button_frame, 
            text=languages[current_lang]["upload_excel"], 
            command=self.upload_excel, 
            fg_color="#e94560", 
            hover_color="#b81d3d",
            corner_radius=8, 
            height=35,
            font=get_font(12, "bold")
        )
        upload_button.pack(side="left", padx=5, fill="x", expand=True)
        
        download_sample_button = ctk.CTkButton(
            button_frame, 
            text=languages[current_lang]["download_sample_excel"], 
            command=self.download_sample_excel, 
            fg_color="#4a5568", 
            hover_color="#2d3748",
            corner_radius=8, 
            height=35,
            font=get_font(12, "bold")
        )
        download_sample_button.pack(side="left", padx=5, fill="x", expand=True)
        
        # Modern styled treeview with reduced height
        self.numbers_tree = ttk.Treeview(
            left_frame, 
            columns=("number", "name"), 
            show="headings", 
            style="Custom.Treeview", 
            height=15
        )
        self.numbers_tree.heading("number", text=languages[current_lang]["number_column"])
        self.numbers_tree.heading("name", text=languages[current_lang]["name_column"])
        self.numbers_tree.column("number", width=120)
        self.numbers_tree.column("name", width=120)
        self.numbers_tree.pack(fill="both", expand=True, padx=15, pady=10)
        
        # Center Panel - Messages with reduced width
        center_frame = ctk.CTkFrame(main_frame, corner_radius=15, fg_color="#2d3748")
        center_frame.grid(row=0, column=1, sticky="nswe", padx=10, pady=10)
        
        # Modern styled message label with reduced font size
        message_label = ctk.CTkLabel(
            center_frame, 
            text=languages[current_lang]["message_section"], 
            font=get_font(20, "bold"), 
            text_color="#e94560"
        )
        message_label.pack(padx=15, pady=15)
        
        # Modern styled notebook with reduced size
        self.message_notebook = ctk.CTkTabview(
            center_frame, 
            width=500, 
            height=300,
            fg_color="#1f2937",
            corner_radius=8
        )
        self.message_notebook.pack(padx=15, pady=10, fill="both", expand=True)
        
        # Style the tabs with reduced size
        self.message_notebook._tabview_button_fg_color = "#4a5568"
        self.message_notebook._tabview_button_hover_color = "#2d3748"
        self.message_notebook._tabview_button_selected_color = "#e94560"
        self.message_notebook._tabview_button_height = 35
        self.message_notebook._tabview_button_font = get_font(12, "bold")
        
        self.textboxes = []
        for i in range(1, 6):
            tab_name = f"{languages[current_lang]['message_tab']} {i}"
            tab = self.message_notebook.add(tab_name)
            textbox = ctk.CTkTextbox(
                tab, 
                width=480, 
                height=250, 
                wrap="word", 
                font=get_font(12),
                fg_color="#1f2937",
                text_color="#e2e8f0"
            )
            textbox.pack(fill="both", expand=True, padx=10, pady=10)
            self.textboxes.append(textbox)
        
        # Translation buttons with modern styling and increased size
        translate_frame = ctk.CTkFrame(center_frame, fg_color="#2d3748")
        translate_frame.pack(padx=20, pady=15, fill="x")
        
        # Define translation text based on current language
        if current_lang == "Persian":
            translate_arabic_text = "ترجمه همه پیام‌ها به عربی"
            translate_english_text = "ترجمه همه پیام‌ها به انگلیسی"
        elif current_lang == "Arabic":
            translate_arabic_text = "ترجمة كل الرسائل إلى العربية"
            translate_english_text = "ترجمة كل الرسائل إلى الإنجليزية"
        else:
            translate_arabic_text = "Translate all messages to Arabic"
            translate_english_text = "Translate all messages to English"
        
        translate_arabic_group_button = ctk.CTkButton(
            translate_frame, 
            text=translate_arabic_text, 
            command=self.translate_group_to_arabic, 
            fg_color="#4a5568", 
            hover_color="#2d3748",
            corner_radius=10, 
            height=40,
            font=get_font(14, "bold")
        )
        translate_arabic_group_button.pack(side="left", padx=5, fill="x", expand=True)
        
        translate_english_group_button = ctk.CTkButton(
            translate_frame, 
            text=translate_english_text, 
            command=self.translate_group_to_english, 
            fg_color="#4a5568", 
            hover_color="#2d3748",
            corner_radius=10, 
            height=40,
            font=get_font(14, "bold")
        )
        translate_english_group_button.pack(side="left", padx=5, fill="x", expand=True)
        
        # Polls and buttons section with modern styling and increased spacing
        polls_buttons_frame = ctk.CTkFrame(center_frame, fg_color="#2d3748")
        polls_buttons_frame.pack(padx=20, pady=15, fill="x")
        
        polls_label = ctk.CTkLabel(
            polls_buttons_frame, 
            text=languages[current_lang]["polls"], 
            font=get_font(18, "bold"), 
            text_color="#e94560"
        )
        polls_label.pack(side="left", padx=15)
        
        add_poll_button = ctk.CTkButton(
            polls_buttons_frame, 
            text=languages[current_lang]["add_poll"], 
            width=150, 
            height=40,
            command=self.open_poll_window, 
            fg_color="#e94560", 
            hover_color="#b81d3d",
            corner_radius=10, 
            font=get_font(14, "bold")
        )
        add_poll_button.pack(side="left", padx=10)
        
        buttons_label = ctk.CTkLabel(
            polls_buttons_frame, 
            text=languages[current_lang]["buttons"], 
            font=get_font(18, "bold"), 
            text_color="#e94560"
        )
        buttons_label.pack(side="left", padx=15)
        
        add_button_button = ctk.CTkButton(
            polls_buttons_frame, 
            text=languages[current_lang]["add_button"], 
            width=150, 
            height=40,
            command=self.placeholder_action, 
            fg_color="#e94560", 
            hover_color="#b81d3d",
            corner_radius=10, 
            font=get_font(14, "bold")
        )
        add_button_button.pack(side="left", padx=10)
        
        # Right Panel - Attachments with increased width
        right_frame = ctk.CTkFrame(main_frame, corner_radius=15, fg_color="#2d3748", width=350)
        right_frame.grid(row=0, column=2, sticky="nswe", padx=15, pady=15)
        
        attachments_label = ctk.CTkLabel(
            right_frame, 
            text=languages[current_lang]["attachments"], 
            font=get_font(24, "bold"), 
            text_color="#e94560"
        )
        attachments_label.pack(padx=20, pady=20)
        
        # Modern styled add file button with increased size
        add_file_button = ctk.CTkButton(
            right_frame, 
            text=languages[current_lang]["add_file"], 
            command=self.add_attachment, 
            fg_color="#e94560", 
            hover_color="#b81d3d",
            corner_radius=10, 
            height=40,
            font=get_font(14, "bold")
        )
        add_file_button.pack(padx=20, pady=15)
        
        # Modern styled attachment type label and menu with increased size
        attachment_type_label = ctk.CTkLabel(
            right_frame, 
            text=languages[current_lang]["attachment_type"], 
            font=get_font(16, "bold"), 
            text_color="#e2e8f0"
        )
        attachment_type_label.pack(padx=20, pady=(15, 0))
        
        attachment_type_menu = ctk.CTkOptionMenu(
            right_frame, 
            variable=self.attachment_type, 
            values=[languages[current_lang]["image"], languages[current_lang]["sticker"], languages[current_lang]["mp3"], languages[current_lang]["video"]], 
            width=150, 
            height=35,
            font=get_font(14),
            fg_color="#1f2937",
            button_color="#e94560",
            button_hover_color="#b81d3d"
        )
        attachment_type_menu.set(languages[current_lang]["image"])
        attachment_type_menu.pack(padx=20, pady=10)
        
        # Modern styled attachments tree with increased height
        self.attachments_tree = ttk.Treeview(
            right_frame, 
            columns=("file",), 
            show="headings", 
            style="Custom.Treeview", 
            height=20
        )
        self.attachments_tree.heading("file", text="File")
        self.attachments_tree.column("file", width=200)
        self.attachments_tree.pack(fill="both", expand=True, padx=20, pady=15)
        
        # Bottom Panel - Settings with increased height
        bottom_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="#16213e")
        bottom_frame.pack(side="bottom", fill="x", padx=20, pady=20)
        
        custom_frame = ctk.CTkFrame(bottom_frame, corner_radius=0, fg_color="#16213e")
        custom_frame.pack(fill="x", padx=20, pady=20)
        
        # Delay settings with modern styling and increased spacing
        delay_frame = ctk.CTkFrame(custom_frame, corner_radius=15, fg_color="#2d3748")
        delay_frame.pack(side="left", padx=15, pady=10)
        
        delay_label = ctk.CTkLabel(
            delay_frame, 
            text=languages[current_lang]["delay_settings"], 
            font=get_font(18, "bold"), 
            text_color="#e94560"
        )
        delay_label.grid(row=0, column=0, columnspan=5, padx=15, pady=15)
        
        ctk.CTkLabel(
            delay_frame, 
            text=languages[current_lang]["wait_label"], 
            font=get_font(14), 
            text_color="#e2e8f0"
        ).grid(row=1, column=0, padx=10, pady=10)
        
        delay_spin = SpinboxAlternative(
            delay_frame, 
            from_=1, 
            to=9999, 
            width=80, 
            textvariable=self.delay_seconds,
            fg_color="#1f2937",
            button_color="#e94560",
            button_hover_color="#b81d3d"
        )
        delay_spin.grid(row=1, column=1, padx=10, pady=10)
        
        ctk.CTkLabel(
            delay_frame, 
            text=languages[current_lang]["seconds_after_every"], 
            font=get_font(14), 
            text_color="#e2e8f0"
        ).grid(row=1, column=2, padx=10, pady=10)
        
        msgs_spin = SpinboxAlternative(
            delay_frame, 
            from_=1, 
            to=9999, 
            width=80, 
            textvariable=self.delay_every_msgs,
            fg_color="#1f2937",
            button_color="#e94560",
            button_hover_color="#b81d3d"
        )
        msgs_spin.grid(row=1, column=3, padx=10, pady=10)
        
        ctk.CTkLabel(
            delay_frame, 
            text=languages[current_lang]["messages_label"], 
            font=get_font(14), 
            text_color="#e2e8f0"
        ).grid(row=1, column=4, padx=10, pady=10)
        
        # Send mode with modern styling and increased spacing
        send_mode_frame = ctk.CTkFrame(custom_frame, corner_radius=15, fg_color="#2d3748")
        send_mode_frame.pack(side="left", padx=15, pady=10)
        
        send_mode_label = ctk.CTkLabel(
            send_mode_frame, 
            text=languages[current_lang]["send_mode"], 
            font=get_font(18, "bold"), 
            text_color="#e94560"
        )
        send_mode_label.grid(row=0, column=0, columnspan=2, padx=15, pady=10)
        
        manual_radio = ctk.CTkRadioButton(
            send_mode_frame, 
            text=languages[current_lang]["manual"], 
            variable=self.send_mode_var, 
            value="Manual", 
            font=get_font(14), 
            text_color="#e2e8f0",
            fg_color="#e94560",
            hover_color="#b81d3d"
        )
        manual_radio.grid(row=1, column=0, padx=10, pady=10)
        
        auto_radio = ctk.CTkRadioButton(
            send_mode_frame, 
            text=languages[current_lang]["auto"], 
            variable=self.send_mode_var, 
            value="Auto", 
            font=get_font(14), 
            text_color="#e2e8f0",
            fg_color="#e94560",
            hover_color="#b81d3d"
        )
        auto_radio.grid(row=1, column=1, padx=10, pady=10)
        
        # Action buttons with modern styling and increased size
        group_send_button = ctk.CTkButton(
            custom_frame, 
            text=languages[current_lang]["group_send"], 
            command=self.open_process_window, 
            fg_color="#e94560", 
            hover_color="#b81d3d",
            corner_radius=10, 
            height=40,
            font=get_font(14, "bold")
        )
        group_send_button.pack(side="left", padx=15, pady=10)
        
        single_send_button = ctk.CTkButton(
            custom_frame, 
            text=languages[current_lang]["single_send"], 
            command=self.open_single_message_window, 
            fg_color="#e94560", 
            hover_color="#b81d3d",
            corner_radius=10, 
            height=40,
            font=get_font(14, "bold")
        )
        single_send_button.pack(side="left", padx=15, pady=10)
        
        exit_button = ctk.CTkButton(
            custom_frame, 
            text=languages[current_lang]["exit_button"], 
            command=self.on_close, 
            fg_color="#ef4444", 
            hover_color="#dc2626",
            corner_radius=10, 
            height=40,
            font=get_font(14, "bold")
        )
        exit_button.pack(side="right", padx=15, pady=10)
        
        # Update Treeview style with modern colors
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Custom.Treeview", 
            background="#1f2937", 
            foreground="#e2e8f0", 
            fieldbackground="#1f2937", 
            font=get_ttk_font()
        )
        style.configure(
            "Custom.Treeview.Heading", 
            background="#2d3748", 
            foreground="#e94560", 
            font=get_ttk_heading_font()
        )
        style.map("Custom.Treeview", background=[("selected", "#4a5568")])
    
    def change_language(self, lang_choice):
        global current_lang
        if lang_choice in languages:
            current_lang = lang_choice
            self.update_texts()
    
    def change_appearance_mode(self, mode):
        ctk.set_appearance_mode(mode.lower())
    
    def update_texts(self):
        self.title(languages[current_lang]["title"])
        self.destroy()
        new_app = WhatsAppMarketingApp()
        new_app.mainloop()
    
    def upload_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")], title=languages[current_lang]["upload_excel"])
        if not file_path:
            return
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            self.numbers_list.clear()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                number = str(row[0]).strip() if row[0] else ""
                name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if number:
                    corrected_number = self.correct_phone_number(number)
                    self.numbers_list.append({"number": corrected_number, "name": name})
            self.refresh_numbers_tree()
        except Exception as e:
            messagebox.showerror(languages[current_lang]["title"], f"Error reading Excel: {e}")
    
    def refresh_numbers_tree(self):
        for item in self.numbers_tree.get_children():
            self.numbers_tree.delete(item)
        for entry in self.numbers_list:
            self.numbers_tree.insert("", tk.END, values=(entry["number"], entry["name"]))
    
    def download_sample_excel(self):
        sample_path = "sample_numbers.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "Number"
        sheet["B1"] = "Name"
        sheet['A2'] = "9358883639"
        sheet['B2'] = "Mani agah"
        workbook.save(sample_path)
        messagebox.showinfo(languages[current_lang]["title"], f"Sample Excel saved as {sample_path}")
    
    def add_attachment(self):
        file_path = filedialog.askopenfilename(title=languages[current_lang]["add_file"])
        if file_path:
            self.attachments_list.append(file_path)
            self.refresh_attachments_tree()
    
    def refresh_attachments_tree(self):
        for item in self.attachments_tree.get_children():
            self.attachments_tree.delete(item)
        for att in self.attachments_list:
            filename = os.path.basename(att)
            self.attachments_tree.insert("", tk.END, values=(filename,))
    
    def open_poll_window(self):
        PollWindow(self)
    
    def open_process_window(self):
        messages = []
        for textbox in self.textboxes:
            content = textbox.get("1.0", "end").strip()
            if content:
                messages.append(content)
        if not self.numbers_list or (not messages and not self.attachments_list and not self.polls_list):
            messagebox.showerror(languages[current_lang]["title"], languages[current_lang]["error_empty"])
            return
        send_window = SendProcessWindow(self, self.numbers_list, messages, self.attachments_list, self.polls_list, self.delay_seconds.get(), self.delay_every_msgs.get(), self.send_mode_var.get())
        if not self.driver:
            self.driver = send_window.driver
        else:
            send_window.driver = self.driver
    
    def open_single_message_window(self):
        SingleMessageInputWindow(self)
    
    def placeholder_action(self):
        messagebox.showinfo("Info", "This feature is just a placeholder. You can implement your own logic.")
    
    def translate_group_to_arabic(self):
        for textbox in self.textboxes:
            original_text = textbox.get("1.0", "end").strip()
            if original_text:
                translated = translate_text(original_text, "ar")
                if translated:
                    textbox.delete("1.0", "end")
                    textbox.insert("end", translated)
                else:
                    messagebox.showerror("Error", "Translation to Arabic failed.")
    
    def translate_group_to_english(self):
        for textbox in self.textboxes:
            original_text = textbox.get("1.0", "end").strip()
            if original_text:
                translated = translate_text(original_text, "en")
                if translated:
                    textbox.delete("1.0", "end")
                    textbox.insert("end", translated)
                else:
                    messagebox.showerror("Error", "Translation to English failed.")
    
    def correct_phone_number(self, phone):
        # Remove any non-digit characters
        phone = ''.join(filter(str.isdigit, str(phone)))
        
        # If the number starts with 0, replace it with 98 (Iran's country code)
        if phone.startswith('0'):
            phone = '98' + phone[1:]
        
        # If the number doesn't start with a country code, add 98 (Iran's country code)
        if not phone.startswith('98'):
            phone = '98' + phone
        
        # Ensure the number is at least 12 digits (country code + number)
        if len(phone) < 12:
            raise ValueError("Phone number is too short")
        
        return phone
    
    def on_close(self):
        if self.driver:
            self.driver.quit()
        self.destroy()

if __name__ == "__main__":
    app = WhatsAppMarketingApp()
    app.mainloop()
