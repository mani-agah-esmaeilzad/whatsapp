import os
import time
import threading
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import customtkinter as ctk
import openpyxl

# Selenium Imports
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ============================
# زبان‌ها و ترجمه‌ها
# ============================
languages = {
    "English": {
        "title": "WhatsApp Marketing",
        "send_button": "Send",
        "connection_status_default": "Connection Status: Unknown",
        "error_connection": "Connection error. Please check your internet or WhatsApp login.",
        "error_empty": "No phone, message, or poll to send.",
        "success_sent": "All messages have been sent!",
        "error_sending": "Error sending message:",
        "exit_button": "Exit",
        "upload_excel": "Upload Excel",
        "download_sample_excel": "Download Sample Excel",
        "target_label": "Targets",
        "number_column": "Number",
        "name_column": "Name",
        "message_section": "Message Section",
        "message_tab": "Message",
        "polls": "Polls",
        "add_poll": "Add Poll",
        "buttons": "Buttons",
        "add_button": "Add Button",
        "attachments": "Attachments",
        "add_file": "Add File",
        "delay_settings": "Delay Settings",
        "wait_label": "Wait",
        "seconds_after_every": "seconds after every",
        "messages_label": "messages",
        "send_mode": "Send Mode",
        "manual": "Manual",
        "auto": "Auto",
        "group_send": "Group Send",
        "single_send": "Single Send",
        "phone_label": "Phone:",
        "single_message_label": "Single Message:",
        "language_option": "English"
    },
    "Persian": {
        "title": "واتساپ مارکتینگ",
        "send_button": "ارسال",
        "connection_status_default": "وضعیت اتصال: نامشخص",
        "error_connection": "خطا در اتصال. لطفاً اینترنت یا ورود به واتساپ را بررسی کنید.",
        "error_empty": "شماره، پیام یا نظرسنجی برای ارسال وجود ندارد.",
        "success_sent": "تمام پیام‌ها ارسال شدند!",
        "error_sending": "خطا در ارسال پیام:",
        "exit_button": "خروج",
        "upload_excel": "بارگذاری اکسل",
        "download_sample_excel": "دانلود نمونه اکسل",
        "target_label": "مخاطبان",
        "number_column": "شماره",
        "name_column": "نام",
        "message_section": "بخش پیام",
        "message_tab": "پیام",
        "polls": "نظرسنجی‌ها",
        "add_poll": "افزودن نظرسنجی",
        "buttons": "دکمه‌ها",
        "add_button": "افزودن دکمه",
        "attachments": "پیوست‌ها",
        "add_file": "افزودن فایل",
        "delay_settings": "تنظیمات تأخیر",
        "wait_label": "انتظار",
        "seconds_after_every": "ثانیه بعد از هر",
        "messages_label": "پیام",
        "send_mode": "وضعیت ارسال",
        "manual": "دستی",
        "auto": "خودکار",
        "group_send": "ارسال گروهی",
        "single_send": "ارسال تکی",
        "phone_label": "شماره:",
        "single_message_label": "پیام تکی:",
        "language_option": "فارسی"
    },
    "Arabic": {
        "title": "تسويق واتساب",
        "send_button": "إرسال",
        "connection_status_default": "حالة الاتصال: غير محددة",
        "error_connection": "خطأ في الاتصال. يرجى التحقق من الإنترنت أو تسجيل الدخول إلى واتساپ.",
        "error_empty": "لا يوجد رقم أو رسالة أو استطلاع للإرسال.",
        "success_sent": "تم إرسال جميع الرسائل!",
        "error_sending": "خطأ في إرسال الرسالة:",
        "exit_button": "خروج",
        "upload_excel": "تحميل إكسل",
        "download_sample_excel": "تحميل عينة إكسل",
        "target_label": "المستهدفون",
        "number_column": "رقم",
        "name_column": "اسم",
        "message_section": "قسم الرسالة",
        "message_tab": "رسالة",
        "polls": "الاستطلاعات",
        "add_poll": "إضافة استطلاع",
        "buttons": "الأزرار",
        "add_button": "إضافة زر",
        "attachments": "المرفقات",
        "add_file": "إضافة ملف",
        "delay_settings": "إعدادات التأخير",
        "wait_label": "انتظار",
        "seconds_after_every": "ثانية بعد كل",
        "messages_label": "رسالة",
        "send_mode": "وضع الإرسال",
        "manual": "يدوي",
        "auto": "تلقائي",
        "group_send": "إرسال جماعي",
        "single_send": "إرسال فردي",
        "phone_label": "رقم الهاتف:",
        "single_message_label": "رسالة فردية:",
        "language_option": "العربية"
    }
}

current_lang = "English"

def get_qr_confirmation_message():
    if current_lang == "Persian":
        return "آیا بارکد را اسکن کردید؟"
    elif current_lang == "Arabic":
        return "هل قمت بمسح رمز الاستجابة السريعة؟"
    else:
        return "Did you scan the QR code?"

def get_font(size, weight="normal"):
    if current_lang == "Persian":
        return ctk.CTkFont(family="vazir", size=size, weight=weight)
    elif current_lang == "Arabic":
        return ctk.CTkFont(family="Arial", size=size, weight=weight)
    else:
        return ctk.CTkFont(family="Segoe UI", size=size, weight=weight)

def get_ttk_font():
    if current_lang == "Persian":
        return ("vazir", 12)
    elif current_lang == "Arabic":
        return ("Arial", 12)
    else:
        return ("Segoe UI", 12)

def get_ttk_heading_font():
    if current_lang == "Persian":
        return ("vazir", 12, "bold")
    elif current_lang == "Arabic":
        return ("Arial", 12, "bold")
    else:
        return ("Segoe UI", 12, "bold")

# ============================
# ویجت جایگزین Spinbox
# ============================
class SpinboxAlternative(ctk.CTkFrame):
    def __init__(self, parent, from_=1, to=9999, width=60, textvariable=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.from_ = from_
        self.to = to
        if textvariable is None:
            self.var = tk.IntVar(value=self.from_)
        else:
            self.var = textvariable
        self.btn_minus = ctk.CTkButton(self, text="-", width=30, command=self.decrement, corner_radius=5)
        self.btn_minus.grid(row=0, column=0, padx=2)
        self.entry = ctk.CTkEntry(self, width=width, textvariable=self.var)
        self.entry.grid(row=0, column=1, padx=2)
        self.btn_plus = ctk.CTkButton(self, text="+", width=30, command=self.increment, corner_radius=5)
        self.btn_plus.grid(row=0, column=2, padx=2)
    
    def increment(self):
        try:
            value = int(self.var.get())
        except ValueError:
            value = self.from_
        if value < self.to:
            value += 1
        self.var.set(value)
    
    def decrement(self):
        try:
            value = int(self.var.get())
        except ValueError:
            value = self.from_
        if value > self.from_:
            value -= 1
        self.var.set(value)

# ============================
# پنجره افزودن نظرسنجی
# ============================
class PollWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        if current_lang == "Persian":
            self.title("افزودن نظرسنجی")
        elif current_lang == "Arabic":
            self.title("إضافة استطلاع")
        else:
            self.title("Add Poll")
        self.geometry("400x400")
        question_label = ctk.CTkLabel(self, text=("سوال نظرسنجی:" if current_lang=="Persian" else ("سؤال الاستطلاع:" if current_lang=="Arabic" else "Poll Question:")), font=get_font(14, "bold"))
        question_label.pack(padx=10, pady=10)
        self.question_entry = ctk.CTkEntry(self, font=get_font(14))
        self.question_entry.pack(padx=10, pady=10, fill="x")
        options_label = ctk.CTkLabel(self, text=("گزینه‌های نظرسنجی (هر گزینه در یک خط):" if current_lang=="Persian" else ("خيارات الاستطلاع (سطر لكل خيار):" if current_lang=="Arabic" else "Poll Options (one per line):")), font=get_font(14, "bold"))
        options_label.pack(padx=10, pady=10)
        self.options_textbox = ctk.CTkTextbox(self, font=get_font(14), height=150, wrap="word")
        self.options_textbox.pack(padx=10, pady=10, fill="both", expand=True)
        add_button = ctk.CTkButton(self, text=("افزودن نظرسنجی" if current_lang=="Persian" else ("إضافة استطلاع" if current_lang=="Arabic" else "Add Poll")), command=self.add_poll, fg_color="#0078AA", corner_radius=5, font=get_font(14, "bold"))
        add_button.pack(padx=10, pady=10)
    
    def add_poll(self):
        question = self.question_entry.get().strip()
        options_text = self.options_textbox.get("1.0", "end").strip()
        if not question or not options_text:
            messagebox.showerror("Error" if current_lang=="English" else ("خطا" if current_lang=="Persian" else "خطأ"),
                                 "Please enter both question and options." if current_lang=="English" else ("لطفاً سوال و گزینه‌ها را وارد کنید." if current_lang=="Persian" else "يرجى إدخال السؤال والخيارات."))
            return
        options = [opt.strip() for opt in options_text.splitlines() if opt.strip()]
        if len(options) < 2:
            messagebox.showerror("Error" if current_lang=="English" else ("خطا" if current_lang=="Persian" else "خطأ"),
                                 "Please enter at least two options." if current_lang=="English" else ("لطفاً حداقل دو گزینه وارد کنید." if current_lang=="Persian" else "يرجى إدخال خيارين على الأقل."))
            return
        poll_data = {"question": question, "options": options}
        self.master.polls_list.append(poll_data)
        messagebox.showinfo("Info" if current_lang=="English" else ("اطلاع" if current_lang=="Persian" else "معلومات"),
                            "Poll added successfully." if current_lang=="English" else ("نظرسنجی با موفقیت افزوده شد." if current_lang=="Persian" else "تم إضافة الاستطلاع بنجاح."))
        self.destroy()

# ============================
# پنجره فرآیند ارسال پیام
# ============================
class SendProcessWindow(ctk.CTkToplevel):
    def __init__(self, parent, numbers_list, messages, attachments, polls, delay_seconds, delay_every_msgs, mode):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        self.title("Run")
        self.geometry("900x550")
        self.numbers_list = numbers_list
        self.messages = messages
        self.attachments = attachments
        self.polls = polls
        self.delay_seconds = delay_seconds
        self.delay_every_msgs = delay_every_msgs
        self.mode = mode
        self.paused = False
        self.stopped = False
        self.driver = None
        self.scan_confirmed = None
        
        # --- Top Frame: Initiate Connection ---
        top_frame = ctk.CTkFrame(self, fg_color="#1f2833", corner_radius=10)
        top_frame.pack(fill="x", pady=10, padx=10)
        initiate_label = ctk.CTkLabel(top_frame, text="Initiate WhatsApp & Scan QR Code from your mobile", font=get_font(14, "bold"), text_color="white")
        initiate_label.pack(side="left", padx=10, pady=10)
        self.status_label = ctk.CTkLabel(top_frame, text=languages[current_lang]["connection_status_default"], font=get_font(14), text_color="white")
        self.status_label.pack(side="right", padx=10)
        self.initiate_button = ctk.CTkButton(top_frame, text="CLICK TO INITIATE", fg_color="#0078AA", command=self.initiate_connection, corner_radius=5)
        self.initiate_button.pack(side="left", padx=10)
        
        # --- Middle Frame: Control Buttons & Progress ---
        middle_frame = ctk.CTkFrame(self, fg_color="#20232a", corner_radius=10)
        middle_frame.pack(fill="x", pady=10, padx=10)
        self.start_button = ctk.CTkButton(middle_frame, text="START", fg_color="#4CAF50", command=self.start_sending, corner_radius=5, font=get_font(14, "bold"))
        self.start_button.pack(side="left", padx=5)
        self.pause_button = ctk.CTkButton(middle_frame, text="PAUSE", fg_color="#FF9800", command=self.pause_sending, corner_radius=5, font=get_font(14, "bold"))
        self.pause_button.pack(side="left", padx=5)
        self.stop_button = ctk.CTkButton(middle_frame, text="STOP", fg_color="#F44336", command=self.stop_sending, corner_radius=5, font=get_font(14, "bold"))
        self.stop_button.pack(side="left", padx=5)
        self.progress_label = ctk.CTkLabel(middle_frame, text="0% Completed [0/{}]".format(len(self.numbers_list)), font=get_font(14), text_color="white")
        self.progress_label.pack(side="right", padx=10)
        
        # --- Log Frame ---
        log_frame = ctk.CTkFrame(self, fg_color="#20232a", corner_radius=10)
        log_frame.pack(fill="both", expand=True, padx=10, pady=10)
        self.log_tree = ttk.Treeview(log_frame, columns=("chat_name", "status"), show="headings", style="Custom.Treeview")
        self.log_tree.heading("chat_name", text="Chat Name")
        self.log_tree.heading("status", text="Status")
        self.log_tree.column("chat_name", width=300)
        self.log_tree.column("status", width=200)
        self.log_tree.pack(fill="both", expand=True, padx=10, pady=10)
        
        # --- Notes Frame ---
        notes_frame = ctk.CTkFrame(self, fg_color="#1f2833", corner_radius=10)
        notes_frame.pack(fill="x", padx=10, pady=10)
        notes_label = ctk.CTkLabel(notes_frame, text=("Important Notes:\n1) Make sure your WhatsApp Web is logged in.\n2) Keep your phone connected to the internet.\n3) Do not close the browser tab during sending.\n"), font=get_font(12), justify="left", text_color="white")
        notes_label.pack(side="left", padx=10, pady=10)
    
    def initiate_connection(self):
        threading.Thread(target=self._initiate_connection_worker, daemon=True).start()
    
    def _initiate_connection_worker(self):
        try:
            self.driver = webdriver.Chrome()
            self.driver.get("https://web.whatsapp.com")
            self.after(0, lambda: self.status_label.configure(text="Status: Waiting for login"))
            WebDriverWait(self.driver, 60).until(
                EC.presence_of_element_located((By.XPATH, "//div[@id='pane-side']"))
            )
            self.after(0, lambda: self.status_label.configure(text="Status: Connected"))
            self.scan_confirmed = None
            self.after(0, self.prompt_scan_confirmation)
            while self.scan_confirmed is None:
                time.sleep(0.1)
            if self.scan_confirmed:
                self.driver.set_window_position(-2000, 0)
            else:
                self.after(0, lambda: messagebox.showwarning("Warning", "Please scan the QR code before proceeding."))
        except Exception as e:
            self.after(0, lambda: self.status_label.configure(text="Status: Disconnected"))
            self.after(0, lambda: messagebox.showerror("Error", f"{languages[current_lang]['error_connection']} {e}"))
    
    def prompt_scan_confirmation(self):
        result = messagebox.askyesno("QR Code Scanned", get_qr_confirmation_message())
        self.scan_confirmed = result
    
    def stop_sending(self):
        self.stopped = True
    
    def pause_sending(self):
        self.paused = True
    
    def start_sending(self):
        if not self.driver:
            messagebox.showerror("Error", "Please initiate connection first.")
            return
        self.paused = False
        self.stopped = False
        self.start_button.configure(state="disabled")
        threading.Thread(target=self.send_messages, daemon=True).start()
    
    def send_messages(self):
        total = len(self.numbers_list)
        for idx, contact in enumerate(self.numbers_list, start=1):
            if self.stopped:
                break
            while self.paused and not self.stopped:
                time.sleep(0.5)
            phone = contact["number"]
            self.update_log(phone, "Sending...")
            try:
                self.driver.get(f"https://web.whatsapp.com/send?phone={phone}")
                time.sleep(3)
                if self.attachments:
                    try:
                        attach_button = WebDriverWait(self.driver, 30).until(
                            EC.element_to_be_clickable((By.XPATH, "//div[@data-testid='clip'] | //div[@title='Attach']"))
                        )
                        attach_button.click()
                        image_input = WebDriverWait(self.driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, "//input[@type='file']"))
                        )
                        image_input.send_keys(self.attachments[0])
                        time.sleep(5)
                        try:
                            caption_box = WebDriverWait(self.driver, 10).until(
                                EC.visibility_of_element_located((By.XPATH, "//div[@data-testid='media-caption']"))
                            )
                        except Exception:
                            caption_box = WebDriverWait(self.driver, 10).until(
                                EC.visibility_of_element_located((By.XPATH, "//div[@contenteditable='true'][@data-tab='10']"))
                            )
                        caption = " ".join(self.messages) if self.messages else ""
                        if caption:
                            caption_box.send_keys(caption)
                        send_button = WebDriverWait(self.driver, 30).until(
                            EC.element_to_be_clickable((By.XPATH, "//span[@data-icon='send']"))
                        )
                        send_button.click()
                        time.sleep(2)
                    except Exception as e:
                        self.update_log(phone, f"Image send failed: {e}")
                else:
                    message_box = WebDriverWait(self.driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, "//footer//div[@contenteditable='true']"))
                    )
                    message_box.click()
                    time.sleep(0.5)
                    for msg in self.messages:
                        personalized_msg = msg.replace("{NAME}", contact.get("name", "")).replace("{NUMBER}", contact.get("number", ""))
                        message_box.send_keys(personalized_msg)
                        message_box.send_keys(Keys.ENTER)
                        time.sleep(1)
                self.update_log(phone, "Success")
            except Exception as e:
                self.update_log(phone, f"Failed: {e}")
            percent = int((idx / total) * 100)
            self.progress_label.configure(text=f"{percent}% Completed [{idx}/{total}]")
            if (idx % self.delay_every_msgs == 0) and (idx < total):
                time.sleep(self.delay_seconds)
    
    def update_log(self, chat_name, status):
        self.log_tree.insert("", tk.END, values=(chat_name, status))
        self.log_tree.update_idletasks()

# ============================
# پنجره ارسال پیام تکی
# ============================
class SingleMessageInputWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        if current_lang == "Persian":
            self.title("ارسال پیام تکی")
        elif current_lang == "Arabic":
            self.title("إرسال رسالة فردية")
        else:
            self.title("Send Single Message")
        self.geometry("500x400")
        self.attachment_path = None
        
        # ورودی شماره تلفن
        phone_frame = ctk.CTkFrame(self, fg_color="#1f2833", corner_radius=10)
        phone_frame.pack(fill="x", padx=10, pady=10)
        phone_label = ctk.CTkLabel(phone_frame, text=languages[current_lang]["phone_label"], font=get_font(14, "bold"), text_color="white")
        phone_label.pack(side="left", padx=10, pady=10)
        self.phone_entry = ctk.CTkEntry(phone_frame, font=get_font(14))
        self.phone_entry.pack(side="left", fill="x", expand=True, padx=10, pady=10)
        
        # ورودی پیام
        message_frame = ctk.CTkFrame(self, fg_color="#1f2833", corner_radius=10)
        message_frame.pack(fill="both", expand=True, padx=10, pady=10)
        message_label = ctk.CTkLabel(message_frame, text=languages[current_lang]["single_message_label"], font=get_font(14, "bold"), text_color="white")
        message_label.pack(anchor="nw", padx=10, pady=10)
        self.message_textbox = ctk.CTkTextbox(message_frame, font=get_font(14), wrap="word")
        self.message_textbox.pack(fill="both", expand=True, padx=10, pady=10)
        
        # دکمه افزودن پیوست
        attach_button = ctk.CTkButton(self, text=languages[current_lang]["add_file"], command=self.attach_file, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        attach_button.pack(padx=10, pady=5)
        
        # دکمه ارسال
        send_button = ctk.CTkButton(self, text=languages[current_lang]["send_button"], command=self.send_single_message, fg_color="#4CAF50", corner_radius=5, font=get_font(14, "bold"))
        send_button.pack(padx=10, pady=10)
    
    def attach_file(self):
        file_path = filedialog.askopenfilename(title=languages[current_lang]["add_file"], filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.gif")])
        if file_path:
            self.attachment_path = file_path
            messagebox.showinfo("Info", f"Attached: {os.path.basename(file_path)}")
    
    def send_single_message(self):
        phone = self.phone_entry.get().strip()
        message = self.message_textbox.get("1.0", "end").strip()
        if not phone or not message:
            messagebox.showerror("Error", languages[current_lang]["error_empty"])
            return
        numbers_list = [{"number": phone, "name": ""}]
        messages = [message]
        attachments = []
        if self.attachment_path:
            attachments.append(self.attachment_path)
        polls = []
        send_window = SendProcessWindow(self.master, numbers_list, messages, attachments, polls, 0, 1, "Single")
        if self.master.driver:
            send_window.driver = self.master.driver
        self.destroy()
        send_window.start_sending()

# ============================
# کلاس اصلی برنامه
# ============================
class WhatsAppMarketingApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.state("zoomed")
        self.title(languages[current_lang]["title"])
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.numbers_list = []
        self.attachments_list = []
        self.polls_list = []
        self.delay_seconds = tk.IntVar(value=3)
        self.delay_every_msgs = tk.IntVar(value=5)
        self.send_mode_var = tk.StringVar(value="Manual")
        self.textboxes = []
        self.driver = None  # نگهداری session مرورگر
        
        self.create_ui()
        self.apply_fullscreen_style()
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Custom.Treeview", background="#0b0c10", foreground="white", fieldbackground="#0b0c10", font=get_ttk_font())
        style.configure("Custom.Treeview.Heading", background="#1f2833", foreground="white", font=get_ttk_heading_font())
        style.map("Custom.Treeview", background=[("selected", "#3a3f47")])
    
    def apply_fullscreen_style(self):
        self.configure(bg="#2c3e50")
    
    def create_ui(self):
        # --- Top Bar ---
        top_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="#1f2833", height=60)
        top_frame.pack(side="top", fill="x", padx=10, pady=10)
        title_label = ctk.CTkLabel(top_frame, text=languages[current_lang]["title"], font=get_font(22, "bold"), text_color="white")
        title_label.pack(side="left", padx=10)
        options_frame = ctk.CTkFrame(top_frame, corner_radius=0, fg_color="#1f2833")
        options_frame.pack(side="right", padx=10)
        appearance_option_menu = ctk.CTkOptionMenu(options_frame, values=["Light", "Dark"], command=self.change_appearance_mode, width=120, font=get_font(14))
        appearance_option_menu.set("Dark")
        appearance_option_menu.pack(side="left", padx=5)
        language_option_menu = ctk.CTkOptionMenu(options_frame, values=["English", "Persian", "Arabic"], command=self.change_language, width=120, font=get_font(14))
        language_option_menu.set(languages[current_lang]["language_option"])
        language_option_menu.pack(side="left", padx=5)
        
        # --- Main Content ---
        main_frame = ctk.CTkFrame(self, fg_color="#20232a", corner_radius=10)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        main_frame.grid_columnconfigure(1, weight=1)
        main_frame.grid_rowconfigure(0, weight=1)
        
        # Left Panel: مخاطبان
        left_frame = ctk.CTkFrame(main_frame, corner_radius=10, fg_color="#3a3f47", width=300)
        left_frame.grid(row=0, column=0, sticky="nswe", padx=5, pady=5)
        target_label = ctk.CTkLabel(left_frame, text=languages[current_lang]["target_label"], font=get_font(20, "bold"), text_color="white")
        target_label.pack(padx=10, pady=10)
        button_frame = ctk.CTkFrame(left_frame, fg_color="#3a3f47")
        button_frame.pack(padx=10, pady=5, fill="x")
        upload_button = ctk.CTkButton(button_frame, text=languages[current_lang]["upload_excel"], command=self.upload_excel, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        upload_button.pack(side="left", padx=5)
        download_sample_button = ctk.CTkButton(button_frame, text=languages[current_lang]["download_sample_excel"], command=self.download_sample_excel, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        download_sample_button.pack(side="left", padx=5)
        self.numbers_tree = ttk.Treeview(left_frame, columns=("number", "name"), show="headings", style="Custom.Treeview", height=15)
        self.numbers_tree.heading("number", text=languages[current_lang]["number_column"])
        self.numbers_tree.heading("name", text=languages[current_lang]["name_column"])
        self.numbers_tree.column("number", width=120)
        self.numbers_tree.column("name", width=120)
        self.numbers_tree.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Center Panel: پیام‌ها و نظرسنجی
        center_frame = ctk.CTkFrame(main_frame, corner_radius=10, fg_color="#3a3f47")
        center_frame.grid(row=0, column=1, sticky="nswe", padx=5, pady=5)
        message_label = ctk.CTkLabel(center_frame, text=languages[current_lang]["message_section"], font=get_font(20, "bold"), text_color="white")
        message_label.pack(padx=10, pady=10)
        self.message_notebook = ctk.CTkTabview(center_frame, width=600, height=300)
        self.message_notebook.pack(padx=10, pady=5, fill="both", expand=True)
        self.textboxes = []
        for i in range(1, 6):
            tab_name = f"{languages[current_lang]['message_tab']} {i}"
            tab = self.message_notebook.add(tab_name)
            textbox = ctk.CTkTextbox(tab, width=580, height=250, wrap="word", font=get_font(14))
            textbox.pack(fill="both", expand=True, padx=10, pady=10)
            self.textboxes.append(textbox)
        polls_buttons_frame = ctk.CTkFrame(center_frame, fg_color="#3a3f47")
        polls_buttons_frame.pack(padx=10, pady=5, fill="x")
        polls_label = ctk.CTkLabel(polls_buttons_frame, text=languages[current_lang]["polls"], font=get_font(16), text_color="white")
        polls_label.pack(side="left", padx=5)
        add_poll_button = ctk.CTkButton(polls_buttons_frame, text=languages[current_lang]["add_poll"], width=100, command=self.open_poll_window, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        add_poll_button.pack(side="left", padx=5)
        buttons_label = ctk.CTkLabel(polls_buttons_frame, text=languages[current_lang]["buttons"], font=get_font(16), text_color="white")
        buttons_label.pack(side="left", padx=5)
        add_button_button = ctk.CTkButton(polls_buttons_frame, text=languages[current_lang]["add_button"], width=100, command=self.placeholder_action, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        add_button_button.pack(side="left", padx=5)
        
        # Right Panel: پیوست‌ها
        right_frame = ctk.CTkFrame(main_frame, corner_radius=10, fg_color="#3a3f47", width=300)
        right_frame.grid(row=0, column=2, sticky="nswe", padx=5, pady=5)
        attachments_label = ctk.CTkLabel(right_frame, text=languages[current_lang]["attachments"], font=get_font(20, "bold"), text_color="white")
        attachments_label.pack(padx=10, pady=10)
        add_file_button = ctk.CTkButton(right_frame, text=languages[current_lang]["add_file"], command=self.add_attachment, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        add_file_button.pack(padx=10, pady=5)
        self.attachments_tree = ttk.Treeview(right_frame, columns=("file",), show="headings", style="Custom.Treeview", height=15)
        self.attachments_tree.heading("file", text="File")
        self.attachments_tree.column("file", width=150)
        self.attachments_tree.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Bottom Panel: تنظیمات تأخیر و حالت ارسال
        bottom_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="#1f2833")
        bottom_frame.pack(side="bottom", fill="x", padx=10, pady=10)
        custom_frame = ctk.CTkFrame(bottom_frame, corner_radius=0, fg_color="#1f2833")
        custom_frame.pack(fill="x", padx=10, pady=10)
        delay_frame = ctk.CTkFrame(custom_frame, corner_radius=10, fg_color="#3a3f47")
        delay_frame.pack(side="left", padx=10, pady=5)
        delay_label = ctk.CTkLabel(delay_frame, text=languages[current_lang]["delay_settings"], font=get_font(16, "bold"), text_color="white")
        delay_label.grid(row=0, column=0, columnspan=5, padx=5, pady=5)
        ctk.CTkLabel(delay_frame, text=languages[current_lang]["wait_label"], font=get_font(14), text_color="white").grid(row=1, column=0, padx=5, pady=5)
        delay_spin = SpinboxAlternative(delay_frame, from_=1, to=9999, width=60, textvariable=self.delay_seconds)
        delay_spin.grid(row=1, column=1, padx=5, pady=5)
        ctk.CTkLabel(delay_frame, text=languages[current_lang]["seconds_after_every"], font=get_font(14), text_color="white").grid(row=1, column=2, padx=5, pady=5)
        msgs_spin = SpinboxAlternative(delay_frame, from_=1, to=9999, width=60, textvariable=self.delay_every_msgs)
        msgs_spin.grid(row=1, column=3, padx=5, pady=5)
        ctk.CTkLabel(delay_frame, text=languages[current_lang]["messages_label"], font=get_font(14), text_color="white").grid(row=1, column=4, padx=5, pady=5)
        send_mode_frame = ctk.CTkFrame(custom_frame, corner_radius=10, fg_color="#3a3f47")
        send_mode_frame.pack(side="left", padx=10, pady=5)
        send_mode_label = ctk.CTkLabel(send_mode_frame, text=languages[current_lang]["send_mode"], font=get_font(16, "bold"), text_color="white")
        send_mode_label.grid(row=0, column=0, columnspan=2, padx=5, pady=5)
        manual_radio = ctk.CTkRadioButton(send_mode_frame, text=languages[current_lang]["manual"], variable=self.send_mode_var, value="Manual", font=get_font(14), text_color="white")
        manual_radio.grid(row=1, column=0, padx=5, pady=5)
        auto_radio = ctk.CTkRadioButton(send_mode_frame, text=languages[current_lang]["auto"], variable=self.send_mode_var, value="Auto", font=get_font(14), text_color="white")
        auto_radio.grid(row=1, column=1, padx=5, pady=5)
        group_send_button = ctk.CTkButton(custom_frame, text=languages[current_lang]["group_send"], command=self.open_process_window, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        group_send_button.pack(side="left", padx=10, pady=5)
        single_send_button = ctk.CTkButton(custom_frame, text=languages[current_lang]["single_send"], command=self.open_single_message_window, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        single_send_button.pack(side="left", padx=10, pady=5)
        exit_button = ctk.CTkButton(custom_frame, text=languages[current_lang]["exit_button"], command=self.on_close, fg_color="#F44336", corner_radius=5, font=get_font(14))
        exit_button.pack(side="right", padx=10, pady=5)
    
    def change_language(self, lang_choice):
        global current_lang
        if lang_choice in languages:
            current_lang = lang_choice
            self.update_texts()
    
    def change_appearance_mode(self, mode):
        ctk.set_appearance_mode(mode.lower())
    
    def update_texts(self):
        self.title(languages[current_lang]["title"])
        self.destroy()
        new_app = WhatsAppMarketingApp()
        new_app.mainloop()
    
    def upload_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")], title=languages[current_lang]["upload_excel"])
        if not file_path:
            return
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            self.numbers_list.clear()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                number = str(row[0]).strip() if row[0] else ""
                name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if number:
                    self.numbers_list.append({"number": number, "name": name})
            self.refresh_numbers_tree()
        except Exception as e:
            messagebox.showerror(languages[current_lang]["title"], f"Error reading Excel: {e}")
    
    def refresh_numbers_tree(self):
        for item in self.numbers_tree.get_children():
            self.numbers_tree.delete(item)
        for entry in self.numbers_list:
            self.numbers_tree.insert("", tk.END, values=(entry["number"], entry["name"]))
    
    def download_sample_excel(self):
        sample_path = "sample_numbers.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "Number"
        sheet["B1"] = "Name"
        sheet['A2'] = "+1234567890"
        sheet['B2'] = "John Doe"
        workbook.save(sample_path)
        messagebox.showinfo(languages[current_lang]["title"], f"Sample Excel saved as {sample_path}")
    
    def add_attachment(self):
        file_path = filedialog.askopenfilename(title=languages[current_lang]["add_file"])
        if file_path:
            self.attachments_list.append(file_path)
            self.refresh_attachments_tree()
    
    def refresh_attachments_tree(self):
        for item in self.attachments_tree.get_children():
            self.attachments_tree.delete(item)
        for att in self.attachments_list:
            filename = os.path.basename(att)
            self.attachments_tree.insert("", tk.END, values=(filename,))
    
    def open_poll_window(self):
        PollWindow(self)
    
    def open_process_window(self):
        messages = []
        for textbox in self.textboxes:
            content = textbox.get("1.0", "end").strip()
            if content:
                messages.append(content)
        if not self.numbers_list or (not messages and not self.attachments_list and not self.polls_list):
            messagebox.showerror(languages[current_lang]["title"], languages[current_lang]["error_empty"])
            return
        send_window = SendProcessWindow(self, self.numbers_list, messages, self.attachments_list, self.polls_list, self.delay_seconds.get(), self.delay_every_msgs.get(), self.send_mode_var.get())
        if not self.driver:
            self.driver = send_window.driver
        else:
            send_window.driver = self.driver
    
    def open_single_message_window(self):
        SingleMessageInputWindow(self)
    
    def placeholder_action(self):
        messagebox.showinfo("Info", "This feature is just a placeholder. You can implement your own logic.")
    
    def on_close(self):
        if self.driver:
            self.driver.quit()
        self.destroy()

if __name__ == "__main__":
    app = WhatsAppMarketingApp()
    app.mainloop()
