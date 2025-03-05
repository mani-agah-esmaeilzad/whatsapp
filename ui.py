import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import tkinter.ttk as ttk
from windows import PollWindow, SendProcessWindow, SingleMessageWindow
from utils import SpinboxAlternative
from config import languages, current_lang, get_font, get_ttk_font, get_ttk_heading_font
from styles import apply_styles


class WhatsAppMarketingApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.state("zoomed")
        self.bind("<F11>", self.toggle_fullscreen)
        self.title(languages[current_lang]["title"])
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.numbers_list = []
        self.attachments_list = []
        self.polls_list = []
        self.delay_seconds = tk.IntVar(value=3)
        self.delay_every_msgs = tk.IntVar(value=5)
        self.send_mode_var = tk.StringVar(value="Manual")
        self.textboxes = []
        self.create_ui()
        self.apply_fullscreen_style()
        apply_styles()

    def toggle_fullscreen(self, event=None):
        if self.state() == "zoomed":
            self.state("normal")
            self.apply_normal_style()
        else:
            self.state("zoomed")
            self.apply_fullscreen_style()

    def apply_fullscreen_style(self):
        self.configure(bg="#2c3e50")

    def apply_normal_style(self):
        self.configure(bg="#f0f0f0")

    def create_ui(self):
        top_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="#1f2833")
        top_frame.pack(side="top", fill="x", padx=10, pady=10)
        accounts_label = ctk.CTkLabel(top_frame, text="Accounts", font=get_font(22, "bold"), text_color="white")
        accounts_label.pack(side="left", padx=10, pady=10)
        top_right_frame = ctk.CTkFrame(top_frame, corner_radius=0, fg_color="#1f2833")
        top_right_frame.pack(side="right", padx=10, pady=10)
        appearance_option_menu = ctk.CTkOptionMenu(top_right_frame, values=["Light", "Dark"], command=self.change_appearance_mode, width=120, font=get_font(14))
        appearance_option_menu.set("Dark")
        appearance_option_menu.pack(side="left", padx=5)
        language_option_menu = ctk.CTkOptionMenu(top_right_frame, values=["English", "Persian"], command=self.change_language, width=120, font=get_font(14))
        language_option_menu.set(languages[current_lang]["language_option"])
        language_option_menu.pack(side="left", padx=5)

        main_frame = ctk.CTkFrame(self, fg_color="#20232a", corner_radius=10)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        left_frame = ctk.CTkFrame(main_frame, corner_radius=10, fg_color="#3a3f47")
        left_frame.grid(row=0, column=0, sticky="nswe", padx=5, pady=5)
        left_frame.grid_rowconfigure(1, weight=1)
        left_frame.grid_columnconfigure(0, weight=1)
        target_label = ctk.CTkLabel(left_frame, text=languages[current_lang]["target_label"], font=get_font(20, "bold"), text_color="white")
        target_label.grid(row=0, column=0, padx=10, pady=10)
        button_frame = ctk.CTkFrame(left_frame, fg_color="#3a3f47")
        button_frame.grid(row=1, column=0, padx=10, pady=5, sticky="we")
        upload_button = ctk.CTkButton(button_frame, text=languages[current_lang]["upload_excel"], command=self.upload_excel, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        upload_button.pack(side="left", padx=5)
        download_sample_button = ctk.CTkButton(button_frame, text=languages[current_lang]["download_sample_excel"], command=self.download_sample_excel, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        download_sample_button.pack(side="left", padx=5)
        self.numbers_listbox = ctk.CTkFrame(left_frame, fg_color="#3a3f47")
        self.numbers_listbox.grid(row=2, column=0, padx=10, pady=5, sticky="nswe")
        self.numbers_tree = ttk.Treeview(self.numbers_listbox, columns=("number", "name"), show="headings", style="Custom.Treeview", height=15)
        self.numbers_tree.heading("number", text=languages[current_lang]["number_column"])
        self.numbers_tree.heading("name", text=languages[current_lang]["name_column"])
        self.numbers_tree.column("number", width=120)
        self.numbers_tree.column("name", width=120)
        self.numbers_tree.pack(fill="both", expand=True, padx=10, pady=10)

        center_frame = ctk.CTkFrame(main_frame, corner_radius=10, fg_color="#3a3f47")
        center_frame.grid(row=0, column=1, sticky="nswe", padx=5, pady=5)
        center_frame.grid_rowconfigure(1, weight=1)
        center_frame.grid_columnconfigure(0, weight=1)
        message_label = ctk.CTkLabel(center_frame, text=languages[current_lang]["message_section"], font=get_font(20, "bold"), text_color="white")
        message_label.grid(row=0, column=0, padx=10, pady=10)
        self.message_notebook = ctk.CTkTabview(center_frame, width=400, height=300)
        self.message_notebook.grid(row=1, column=0, padx=10, pady=5, sticky="nswe")
        self.textboxes = []
        for i in range(1, 6):
            tab_name = f"{languages[current_lang]['message_tab']} {i}"
            tab = self.message_notebook.add(tab_name)
            textbox = ctk.CTkTextbox(tab, width=380, height=200, wrap="word", font=get_font(14))
            textbox.pack(fill="both", expand=True, padx=10, pady=10)
            self.textboxes.append(textbox)

        translate_frame = ctk.CTkFrame(center_frame, fg_color="#3a3f47")
        translate_frame.grid(row=3, column=0, padx=10, pady=5, sticky="we")
        self.translate_var = tk.BooleanVar(value=False)
        translate_checkbox = ctk.CTkCheckBox(translate_frame, text="Arabic", variable=self.translate_var, command=self.translate_to_arabic, font=get_font(14), text_color="white")
        translate_checkbox.pack(side="left", padx=5)


        translate_frame_english = ctk.CTkFrame(center_frame, fg_color="#3a3f47")
        translate_frame_english.grid(row=3, column=1, padx=10, pady=5, sticky="we")
        self.translate_var = tk.BooleanVar(value=False)
        translate_checkbox_english = ctk.CTkCheckBox(translate_frame_english, text="English", variable=self.translate_var, command=self.translate_to_english, font=get_font(14), text_color="white")
        translate_checkbox_english.pack(side="right", padx=5)

        polls_buttons_frame = ctk.CTkFrame(center_frame, fg_color="#3a3f47")
        polls_buttons_frame.grid(row=2, column=0, padx=10, pady=5, sticky="we")
        polls_label = ctk.CTkLabel(polls_buttons_frame, text=languages[current_lang]["polls"], font=get_font(16), text_color="white")
        polls_label.pack(side="left", padx=5)
        add_poll_button = ctk.CTkButton(polls_buttons_frame, text=languages[current_lang]["add_poll"], width=100, command=self.open_poll_window, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        add_poll_button.pack(side="left", padx=5)
        buttons_label = ctk.CTkLabel(polls_buttons_frame, text=languages[current_lang]["buttons"], font=get_font(16), text_color="white")
        buttons_label.pack(side="left", padx=5)
        add_button_button = ctk.CTkButton(polls_buttons_frame, text=languages[current_lang]["add_button"], width=100, command=self.placeholder_action, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        add_button_button.pack(side="left", padx=5)

        right_frame = ctk.CTkFrame(main_frame, corner_radius=10, fg_color="#3a3f47")
        right_frame.grid(row=0, column=2, sticky="nswe", padx=5, pady=5)
        right_frame.grid_rowconfigure(1, weight=1)
        attachments_label = ctk.CTkLabel(right_frame, text=languages[current_lang]["attachments"], font=get_font(20, "bold"), text_color="white")
        attachments_label.grid(row=0, column=0, padx=10, pady=10)
        add_file_button = ctk.CTkButton(right_frame, text=languages[current_lang]["add_file"], command=self.add_attachment, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        add_file_button.grid(row=0, column=1, padx=5, pady=10)
        self.attachments_listbox = ctk.CTkFrame(right_frame, fg_color="#3a3f47")
        self.attachments_listbox.grid(row=1, column=0, columnspan=2, padx=10, pady=5, sticky="nswe")
        self.attachments_tree = ttk.Treeview(self.attachments_listbox, columns=("file",), show="headings", style="Custom.Treeview", height=15)
        self.attachments_tree.heading("file", text="File")
        self.attachments_tree.column("file", width=150)
        self.attachments_tree.pack(fill="both", expand=True, padx=10, pady=10)

        bottom_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="#1f2833")
        bottom_frame.pack(side="bottom", fill="x", padx=10, pady=10)
        custom_frame = ctk.CTkFrame(bottom_frame, corner_radius=0, fg_color="#1f2833")
        custom_frame.pack(fill="x", padx=10, pady=10)

        delay_frame = ctk.CTkFrame(custom_frame, corner_radius=10, fg_color="#3a3f47")
        delay_frame.pack(side="left", padx=10, pady=5)
        delay_label = ctk.CTkLabel(delay_frame, text=languages[current_lang]["delay_settings"], font=get_font(16, "bold"), text_color="white")
        delay_label.grid(row=0, column=0, columnspan=5, padx=5, pady=5)
        ctk.CTkLabel(delay_frame, text=languages[current_lang]["wait_label"], font=get_font(14), text_color="white").grid(row=1, column=0, padx=5, pady=5)
        delay_spin = SpinboxAlternative(delay_frame, from_=1, to=9999, width=60, textvariable=self.delay_seconds)
        delay_spin.grid(row=1, column=1, padx=5, pady=5)
        ctk.CTkLabel(delay_frame, text=languages[current_lang]["seconds_after_every"], font=get_font(14), text_color="white").grid(row=1, column=2, padx=5, pady=5)
        msgs_spin = SpinboxAlternative(delay_frame, from_=1, to=9999, width=60, textvariable=self.delay_every_msgs)
        msgs_spin.grid(row=1, column=3, padx=5, pady=5)
        ctk.CTkLabel(delay_frame, text=languages[current_lang]["messages_label"], font=get_font(14), text_color="white").grid(row=1, column=4, padx=5, pady=5)

        send_mode_frame = ctk.CTkFrame(custom_frame, corner_radius=10, fg_color="#3a3f47")
        send_mode_frame.pack(side="left", padx=10, pady=5)
        send_mode_label = ctk.CTkLabel(send_mode_frame, text=languages[current_lang]["send_mode"], font=get_font(16, "bold"), text_color="white")
        send_mode_label.grid(row=0, column=0, columnspan=2, padx=5, pady=5)
        manual_radio = ctk.CTkRadioButton(send_mode_frame, text=languages[current_lang]["manual"], variable=self.send_mode_var, value="Manual", font=get_font(14), text_color="white")
        manual_radio.grid(row=1, column=0, padx=5, pady=5)
        auto_radio = ctk.CTkRadioButton(send_mode_frame, text=languages[current_lang]["auto"], variable=self.send_mode_var, value="Auto", font=get_font(14), text_color="white")
        auto_radio.grid(row=1, column=1, padx=5, pady=5)

        group_send_button = ctk.CTkButton(custom_frame, text=languages[current_lang]["group_send"], command=self.open_process_window, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        group_send_button.pack(side="left", padx=10, pady=5)
        single_send_button = ctk.CTkButton(custom_frame, text=languages[current_lang]["single_send"], command=self.open_single_message_window, fg_color="#0078AA", corner_radius=5, font=get_font(14))
        single_send_button.pack(side="left", padx=10, pady=5)
        exit_button = ctk.CTkButton(custom_frame, text=languages[current_lang]["exit_button"], command=self.on_close, fg_color="#F44336", corner_radius=5, font=get_font(14))
        exit_button.pack(side="right", padx=10, pady=5)

        main_frame.grid_columnconfigure(0, weight=1, minsize=200)
        main_frame.grid_columnconfigure(1, weight=3)
        main_frame.grid_columnconfigure(2, weight=1, minsize=200)
        main_frame.grid_rowconfigure(0, weight=1)

    def open_poll_window(self):
        PollWindow(self)

    def change_language(self, lang_choice):
        global current_lang
        if lang_choice in languages:
            current_lang = lang_choice
            self.update_texts()

    def change_appearance_mode(self, mode):
        ctk.set_appearance_mode(mode.lower())

    def update_texts(self):
        self.title(languages[current_lang]["title"])
        self.destroy()
        new_app = WhatsAppMarketingApp()
        new_app.mainloop()

    def upload_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")], title=languages[current_lang]["upload_excel"])
        if not file_path:
            return
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            self.numbers_list.clear()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                number = str(row[0]).strip() if row[0] else ""
                name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if number:
                    self.numbers_list.append({"number": number, "name": name})
            self.refresh_numbers_tree()
        except Exception as e:
            messagebox.showerror(languages[current_lang]["title"], f"Error reading Excel: {e}")

    def refresh_numbers_tree(self):
        for item in self.numbers_tree.get_children():
            self.numbers_tree.delete(item)
        for entry in self.numbers_list:
            self.numbers_tree.insert("", tk.END, values=(entry["number"], entry["name"]))

    def download_sample_excel(self):
        import openpyxl
        sample_path = "sample_numbers.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "Number"
        sheet["B1"] = "Name"
        sheet["A2"] = "+989358883639"
        sheet["B2"] = "maniagah"
        sheet["A3"] = "+989126602987"
        sheet["B3"] = "maniagah"
        sheet["A4"] = "+989160304331"
        sheet["B4"] = "maniagah"
        sheet["A5"] = "+989010522180"
        sheet["B5"] = "maniagah"
        sheet["A6"] = "+989128885103"
        sheet["B6"] = "maniagah"


        workbook.save(sample_path)
        messagebox.showinfo(languages[current_lang]["title"], f"Sample Excel saved as {sample_path}")

    def add_attachment(self):
        file_path = filedialog.askopenfilename(title=languages[current_lang]["add_file"])
        if file_path:
            self.attachments_list.append(file_path)
            self.refresh_attachments_tree()

    def refresh_attachments_tree(self):
        for item in self.attachments_tree.get_children():
            self.attachments_tree.delete(item)
        for att in self.attachments_list:
            import os
            filename = os.path.basename(att)
            self.attachments_tree.insert("", tk.END, values=(filename,))

    def open_process_window(self):
        if not self.numbers_list:
            messagebox.showerror(languages[current_lang]["title"], languages[current_lang]["error_empty"])
            return
        messages = [textbox.get("1.0", "end").strip() for textbox in self.textboxes if textbox.get("1.0", "end").strip()]
        if not messages and not self.attachments_list and not self.polls_list:
            messagebox.showerror(languages[current_lang]["title"], languages[current_lang]["error_empty"])
            return
        SendProcessWindow(self, self.numbers_list, messages, self.attachments_list, self.polls_list, self.delay_seconds.get(), self.delay_every_msgs.get(), self.send_mode_var.get())

    def open_single_message_window(self):
        SingleMessageWindow(self)

    def translate_to_arabic(self):
        from utils import translate_to_arabic
        translate_to_arabic(self.textboxes, self.translate_var)

    def translate_to_english(self):
        from utils import translate_to_english
        translate_to_english(self.textboxes, self.translate_var)

    def placeholder_action(self):
        messagebox.showinfo("Info", "This feature is just a placeholder. You can implement your own logic.")

    def on_close(self):
        self.destroy()